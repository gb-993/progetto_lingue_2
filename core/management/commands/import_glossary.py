# core/management/commands/import_glossary.py

from django.core.management.base import BaseCommand
from django.db import transaction
import os

try:
    from openpyxl import load_workbook
    HAS_XLSX = True
except Exception:
    HAS_XLSX = False

from core.models import Glossary

DATA_DIR = "data"


def _coerce_str(x):
    """Normalizza valori Excel numerici/date in str quando necessario"""
    if x is None:
        return None
    return str(x)


def read_xlsx_by_filename(base_name):
    """
    Legge data/<base>.xlsx se esiste e restituisce list[dict].
    Usa la PRIMA sheet del file; header = prima riga.
    """
    if not HAS_XLSX:
        return [], False

    if not base_name.lower().endswith(".xlsx"):
        base_name = base_name + ".xlsx"

    path = os.path.join(DATA_DIR, base_name)
    if not os.path.exists(path):
        return [], False

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)

    try:
        headers = next(rows_iter)
    except StopIteration:
        return [], True

    headers = [(_coerce_str(h) or "").strip() for h in headers]
    out = []
    for row in rows_iter:
        d = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            val = row[i] if i < len(row) else None
            d[h] = val
        if any(v is not None and str(v).strip() != "" for v in d.values()):
            out.append(d)
    return out, True


class Command(BaseCommand):
    help = "Importa/aggiorna le voci del glossario dal file data/glossary.xlsx"

    def add_arguments(self, parser):
        parser.add_argument(
            '--clear',
            action='store_true',
            help='Elimina tutte le voci esistenti prima di importare',
        )

    @transaction.atomic
    def handle(self, *args, **options):
        if not HAS_XLSX:
            self.stdout.write(
                self.style.ERROR(
                    "openpyxl non installato. Installa con: pip install openpyxl"
                )
            )
            return

        clear = options.get('clear', False)

        if clear:
            count = Glossary.objects.count()
            Glossary.objects.all().delete()
            self.stdout.write(
                self.style.WARNING(f"Eliminate {count} voci esistenti")
            )

        rows, found = read_xlsx_by_filename("glossary")

        if not found:
            self.stdout.write(
                self.style.ERROR(
                    "File data/glossary.xlsx non trovato!"
                )
            )
            return

        if not rows:
            self.stdout.write(
                self.style.WARNING(
                    "File glossary.xlsx trovato ma vuoto o senza dati validi"
                )
            )
            return

        created_count = 0
        updated_count = 0

        for r in rows:
            word = _coerce_str(r.get("word") or r.get("Column1") or "").strip()
            description = _coerce_str(r.get("description") or r.get("Column2") or "").strip()

            if not word:
                continue

            obj, created = Glossary.objects.update_or_create(
                word=word,
                defaults={"description": description},
            )

            if created:
                created_count += 1
            else:
                updated_count += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"✓ Glossario importato: {created_count} nuove voci, {updated_count} aggiornate"
            )
        )

