
from __future__ import annotations


from types import SimpleNamespace
from typing import Any
import os
import tempfile  
import json
import io
import zipfile
import threading          
import logging         
from django.contrib import messages
from django.utils.safestring import mark_safe
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.management import call_command  
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Prefetch, Count
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.translation import gettext as _t
from django.views.decorators.http import require_http_methods, require_POST
from django.http import HttpRequest, HttpResponse, Http404, JsonResponse, FileResponse
import glob
from django.utils.timezone import now
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from django.utils.translation import gettext as _t
from django.urls import reverse
from datetime import datetime, time, date  
from django.utils import timezone   
from .forms import LanguageForm 


from core.models import (
    Language,
    ParameterDef,
    ParameterReviewFlag,
    Question,
    Answer,
    Example,
    Motivation,
    AnswerMotivation,
    QuestionAllowedMotivation,
    LanguageParameter,
    Glossary,
    ParameterChangeLog,
)
try:
    from core.models import LanguageParameterEval  
    HAS_EVAL = True
except Exception:
    LanguageParameterEval = None 
    HAS_EVAL = False  

try:
    from core.models import AnswerStatus  
except Exception:
    class AnswerStatus:
        PENDING = "pending"
        WAITING = "waiting_for_approval"
        APPROVED = "approved"
        REJECTED = "rejected"

from core.models import LanguageReview 

from core.services.dag_eval import run_dag_for_language
from core.services.dag_debug import diagnostics_for_language


# -----------------------
# Helpers & guardrail
# -----------------------
def _is_admin(user: Any) -> bool:
    """Check whether the given user has administrative permissions.

    Args:
        user: User-like object attached to the current request.

    Returns:
        ``True`` if the user role is ``admin`` or the user is staff/superuser,
        otherwise ``False``.
    """
    return (getattr(user, "role", "") == "admin") or bool(user.is_staff) or bool(user.is_superuser)


def _check_language_access(user: Any, lang: Language) -> bool:
    """Validate whether a user can access a language.

    Admin users are always allowed. Non-admin users can access the language
    only when it is assigned through FK or M2M relations.

    Args:
        user: User-like object attached to the current request.
        lang: Language object being requested.

    Returns:
        ``True`` when access is granted, otherwise ``False``.
    """
    if _is_admin(user):
        return True
    if getattr(lang, "assigned_user_id", None) == user.id:
        return True
    try:
        if lang.users.filter(pk=user.pk).exists():
            return True
    except Exception:
        pass
    return False


def _language_status_summary(lang: Language) -> dict[str, Any]:
    """Compute status counters and compatibility overall status for a language.

    Args:
        lang: Language whose answers are analyzed.

    Returns:
        A dictionary with:
            - ``counts``: per-status totals.
            - ``total``: total number of counted answers.
            - ``overall``: derived status string.
    """
    # Una query: GROUP BY status
    rows = (
        Answer.objects
        .filter(language=lang)
        .values("status")
        .annotate(c=Count("id"))
    )

    counts = {
        "pending": 0,
        "waiting_for_approval": 0,
        "approved": 0,
        "rejected": 0,
    }
    total = 0

    for r in rows:
        s = r["status"] or ""
        if s in counts:
            counts[s] += r["c"]
            total += r["c"]

    if total == 0:
        overall = "pending"
    elif counts["approved"] == total:
        overall = "approved"
    elif counts["waiting_for_approval"] == total:
        overall = "waiting_for_approval"
    elif counts["rejected"] > 0 and counts["waiting_for_approval"] == 0:
        overall = "rejected"
    else:
        overall = "pending"

    return {"counts": counts, "total": total, "overall": overall}



def _language_overall_status(lang: Language) -> dict[str, Any]:
    """Compute the language overall status from answer statuses.

    Priority order is: ``WAITING`` > ``APPROVED`` > ``REJECTED`` > ``PENDING``.

    Args:
        lang: Language whose answers are analyzed.

    Returns:
        A dictionary containing the computed ``overall`` status and the
        per-status ``counts`` map.
    """
    rows = (
        Answer.objects
        .filter(language=lang)
        .values("status")
        .annotate(c=Count("id"))
    )

    counts = {
        "pending": 0,
        "waiting": 0,
        "approved": 0,
        "rejected": 0,
    }

    # mappa status DB -> chiavi counts
    key_map = {
        AnswerStatus.PENDING: "pending",
        AnswerStatus.WAITING: "waiting",
        AnswerStatus.APPROVED: "approved",
        AnswerStatus.REJECTED: "rejected",
    }

    seen_status = set()
    for r in rows:
        status = r["status"]
        seen_status.add(status)
        key = key_map.get(status)
        if key:
            counts[key] += r["c"]

    if AnswerStatus.WAITING in seen_status:
        overall = AnswerStatus.WAITING
    elif AnswerStatus.APPROVED in seen_status:
        overall = AnswerStatus.APPROVED
    elif AnswerStatus.REJECTED in seen_status:
        overall = AnswerStatus.REJECTED
    else:
        overall = AnswerStatus.PENDING

    return {"overall": overall, "counts": counts}



def _all_questions_answered(language: Language) -> bool:
    """Return whether all active questions have a yes/no answer.

    Args:
        language: Language to evaluate.

    Returns:
        ``True`` when every active question has a ``yes`` or ``no`` answer for
        the given language; otherwise ``False``.
    """
    active_qids = set(
        Question.objects.filter(parameter__is_active=True).values_list("id", flat=True)
    )
    if not active_qids:
        return False 
    answered_qids = set(
        Answer.objects.filter(
            language=language,
            question_id__in=active_qids,
            response_text__in=["yes", "no"],
        ).values_list("question_id", flat=True)
    )
    return active_qids.issubset(answered_qids)


# -----------------------
# List / CRUD lingua
# -----------------------

from django.db.models import Max, Q
from django.db.models.functions import Lower
from urllib.parse import urlencode
from django.shortcuts import render
from django.contrib.auth.decorators import login_required

@login_required
@require_http_methods(["GET"])
def language_list(request: HttpRequest) -> HttpResponse:
    """Render the language list with filtering and sortable columns."""

    q = (request.GET.get("q") or "").strip()

    # --- Nuovi Filtri ---
    f_lang_top_family = request.GET.get("f_lang_top_family", "").strip()
    f_lang_family = request.GET.get("f_lang_family", "").strip()
    f_lang_hist = request.GET.get("f_lang_hist", "all").strip()
    f_lang_grp = request.GET.get("f_lang_grp", "").strip()

    # Parametri per ordinamento
    sort_key = (request.GET.get("sort") or "name").strip()
    sort_dir = (request.GET.get("dir") or "asc").strip()

    user = request.user
    is_admin = _is_admin(user)

    qs = (
        Language.objects
        .select_related("assigned_user")
        .annotate(last_change=Max("answers__updated_at"))
    )

    if not is_admin:
        qs = qs.filter(Q(assigned_user=user) | Q(users=user))

    # --- Filtro Ricerca Testuale ---
    if q:
        filt = (
            Q(id__icontains=q)
            | Q(name_full__icontains=q)
            | Q(isocode__icontains=q)
            | Q(glottocode__icontains=q)
            | Q(grp__icontains=q)
            | Q(informant__icontains=q)
            | Q(supervisor__icontains=q)
            | Q(family__icontains=q)
            | Q(top_level_family__icontains=q)
            | Q(source__icontains=q)
        )
        if q.lower() in {"hist", "stor", "storica", "storico", "true", "yes"}:
            filt |= Q(historical_language=True)
        if q.lower() in {"false", "no"}:
            filt |= Q(historical_language=False)
        if is_admin:
            filt |= Q(assigned_user__email__icontains=q)
        qs = qs.filter(filt)

    # --- Applicazione Filtri Avanzati ---
    if f_lang_top_family:
        qs = qs.filter(top_level_family=f_lang_top_family)
    if f_lang_family:
        qs = qs.filter(family=f_lang_family)
    if f_lang_grp:
        qs = qs.filter(grp=f_lang_grp)
    
    if f_lang_hist == "yes":
        qs = qs.filter(historical_language=True)
    elif f_lang_hist == "no":
        qs = qs.filter(historical_language=False)

    # --- Popolamento Dropdown 
    opt_top_families = Language.objects.exclude(top_level_family__isnull=True).exclude(top_level_family="").values_list("top_level_family", flat=True).distinct().order_by("top_level_family")
    opt_families = Language.objects.exclude(family__isnull=True).exclude(family="").values_list("family", flat=True).distinct().order_by("family")
    opt_groups = Language.objects.exclude(grp__isnull=True).exclude(grp="").values_list("grp", flat=True).distinct().order_by("grp")

    sort_map = {
        "id": "id",
        "name": "name_full",
        "top": "top_level_family",
        "family": "family",
        "group": "grp",
        "modified": "last_change",
        "lat": "latitude",
        "lon": "longitude",
    }

    active_sort = None
    if sort_key in sort_map:
        active_sort = sort_key
        if sort_key == "name":
            qs = qs.annotate(_name_ci=Lower("name_full"))
            order_field = "_name_ci"
        elif sort_key == "top":
            qs = qs.annotate(_top_ci=Lower("top_level_family"))
            order_field = "_top_ci"
        else:
            order_field = sort_map[sort_key]

        if sort_dir == "desc":
            qs = qs.order_by(f"-{order_field}", "position")
        else:
            qs = qs.order_by(order_field, "position")

    # Genera gli URL per l'ordinamento mantenendo TUTTI i parametri correnti (ricerca + filtri)
    def _toggle_url(column: str) -> str:
        next_dir = "desc" if (active_sort == column and sort_dir == "asc") else "asc"
        get_params = request.GET.copy()
        get_params["sort"] = column
        get_params["dir"] = next_dir
        return "?" + get_params.urlencode()

    sort_urls = {k: _toggle_url(k) for k in sort_map.keys()}



    map_languages = qs.exclude(latitude__isnull=True).exclude(longitude__isnull=True)
    
    map_data = []
    for lang in map_languages:
        try:
            map_data.append({
                'id': lang.id,
                'name': lang.name_full,
                'lat': float(lang.latitude),  
                'lng': float(lang.longitude),
                'family': lang.family if lang.family else 'Unknown',
                'top_level_family': lang.top_level_family if lang.top_level_family else 'Unknown'
            })
        except (ValueError, TypeError):
            # Ignora le lingue le cui coordinate non sono convertibili in float
            continue

    map_data_json = json.dumps(map_data)


    ctx = {
            "languages": qs,
            "page_obj": None,
            "q": q,
            "is_admin": is_admin,
            "sort": active_sort,
            "dir": sort_dir,
            "sort_urls": sort_urls,
            "params": request.GET,
            "opt_top_families": opt_top_families,
            "opt_families": opt_families,
            "opt_groups": opt_groups,
            "map_data_json": map_data_json, 
        }
    return render(request, "languages/list.html", ctx)





@login_required
@require_http_methods(["GET", "POST"])
def language_add(request: HttpRequest) -> HttpResponse:
    """Create a language entry.

    Args:
        request: Current authenticated HTTP request.

    Returns:
        Form page on GET/validation error, or redirect to ``language_list`` on
        successful creation.
    """
    from .forms import LanguageForm  
    if request.method == "POST":
        form = LanguageForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, _t("Language created."))
            return redirect("language_list")
    else:
        form = LanguageForm()
    return render(request, "languages/add.html", {"page_title": "Add language", "form": form})


@login_required
@require_http_methods(["GET", "POST"])
def language_edit(request: HttpRequest, lang_id: str) -> HttpResponse:
    """Edit an existing language entry.

    Args:
        request: Current authenticated HTTP request.
        lang_id: Primary key of the language to edit.

    Returns:
        Form page on GET/validation error, or redirect to ``language_list`` on
        successful update.
    """
    lang = get_object_or_404(Language, pk=lang_id)
    if request.method == "POST":
        form = LanguageForm(request.POST, instance=lang)
        if form.is_valid():
            form.save()
            messages.success(request, _t("Language updated."))
            return redirect("language_list")
    else:
        form = LanguageForm(instance=lang)
    return render(request, "languages/edit.html", {"page_title": "Edit language", "form": form, "language": lang})



@login_required
@require_POST
@transaction.atomic
def language_delete(request: HttpRequest, lang_id: str) -> HttpResponse:
    """Delete a language after admin password confirmation.

    Args:
        request: Current authenticated HTTP request.
        lang_id: Primary key of the language to delete.

    Returns:
        Redirect response to the appropriate page after authorization and
        deletion checks.
    """
    if not _is_admin(request.user):
        messages.error(request, _t("You are not allowed to perform this action."))
        return redirect("language_list")

    lang = get_object_or_404(Language, pk=lang_id)

    pwd = request.POST.get("admin_password") or ""
    if not request.user.check_password(pwd):
        messages.error(request, _t("Incorrect password. Deletion aborted."))
        return redirect("language_edit", lang_id=lang_id)

    lang.delete()
    messages.success(request, _t(f"Language “{lang_id}” and related data deleted."))
    return redirect("language_list")



# -----------------------
# Pagina data/compilazione
# -----------------------
@login_required
def language_data(request: HttpRequest, lang_id: str) -> HttpResponse:
    """Render the language compilation page with parameters and answers.

    Args:
        request: Current authenticated HTTP request.
        lang_id: Primary key of the language being compiled.

    Returns:
        Rendered data compilation page, or redirect when access is denied.
    """
    user = request.user
    is_admin = _is_admin(user)
    lang = get_object_or_404(Language, pk=lang_id)
    if not _check_language_access(user, lang):
        messages.error(request, _t("You don't have access to this language."))
        return redirect("language_list")

    questions_qs = Question.objects.order_by("id")
    through_qs = QuestionAllowedMotivation.objects.select_related("motivation").order_by("position", "id")
    parameters = (
        ParameterDef.objects.filter(is_active=True)
        .order_by("position")
        .prefetch_related(
            Prefetch("questions", queryset=questions_qs),
            Prefetch("questions__allowed_motivation_links", queryset=through_qs, to_attr="pref_links"),
        )
    )

    answers_qs = (
        Answer.objects.filter(language=lang)
        .select_related("question")
        .prefetch_related("answer_motivations__motivation", "examples")
    )
    answers_by_qid = {a.question_id: a for a in answers_qs}

    # Ora guarda se il parametro è rosso, a prescindere da chi lo ha flaggato
    flagged_pids = set(
        ParameterReviewFlag.objects.filter(language=lang, flag=True).values_list("parameter_id", flat=True)
    )

    active_q_total = 0
    active_q_answered = 0

    for p in parameters:
        total = 0
        answered = 0
        for q in p.questions.all():
            total += 1
            active_q_total += 1
            links = getattr(q, "pref_links", [])
            q.allowed_motivations_list = [l.motivation for l in links] if links else []
            a = answers_by_qid.get(q.id)
            if a:
                q.ans = SimpleNamespace(
                    response_text=(a.response_text or ""),
                    comments=(a.comments or ""),
                    motivation_ids=[am.motivation_id for am in a.answer_motivations.all()],
                    examples=list(a.examples.all()),
                    answer_id=a.id,
                )
                if a.response_text in ("yes", "no"):
                    answered += 1
                    active_q_answered += 1
            else:
                q.ans = SimpleNamespace(response_text="", comments="", motivation_ids=[], examples=[], answer_id=None)

        is_touched = answered > 0
        is_complete = (total > 0 and answered == total)
        is_flagged = p.id in flagged_pids

        if not is_touched:
            p.status = "untouched"
            p.bg_color = "transparent"
            p.fg_color = "inherit"
        elif is_complete and not is_flagged:
            p.status = "ok"
            p.bg_color = "#e6f7e9"
            p.fg_color = "#0f5132"
        else:
            p.status = "red"
            p.bg_color = "#ffe8e8"
            p.fg_color = "#842029"

    all_answered = (active_q_total > 0 and active_q_answered == active_q_total)
    lang_status = _language_overall_status(lang)
    last_reject = (
        LanguageReview.objects.filter(language=lang, decision="reject").order_by("-created_at").first()
    )

    ctx = {
        "language": lang,
        "parameters": parameters,
        "is_admin": is_admin,
        "all_answered": all_answered,
        "lang_status": lang_status,
        "last_reject": last_reject,
    }
    return render(request, "languages/data.html", ctx)


@login_required
@require_http_methods(["POST"])
@transaction.atomic
def parameter_save(request: HttpRequest, lang_id: str, param_id: str) -> HttpResponse:
    """Save all answers for a parameter and manage related review flags.

    Args:
        request: Current authenticated HTTP request.
        lang_id: Primary key of the target language.
        param_id: Primary key of the parameter being saved.

    Returns:
        Redirect response to the language data page (with parameter anchor) or
        to the language list when the language no longer exists.
    """

    try:
        lang = Language.objects.select_for_update().get(pk=lang_id)
    except Language.DoesNotExist:
        messages.warning(request, _t("This language was deleted while you were saving. Your changes were not saved."))
        return redirect("language_list")
    if not _check_language_access(request.user, lang):
        messages.error(request, _t("You don't have access to this language."))
        return redirect("language_list")

    param = get_object_or_404(ParameterDef, pk=param_id, is_active=True)
    questions = list(param.questions.all())

    saved_count = 0
    FIELDS = {"number", "textarea", "transliteration", "gloss", "translation", "reference"}

    for q in questions:
        response_text = (request.POST.get(f"resp_{q.id}") or "").strip().lower()
        comments = (request.POST.get(f"com_{q.id}") or "").strip()

        if response_text not in ("yes", "no"):
            continue

        # Lock dell'Answer se esiste (evita corse su update concorrenti); se non esiste, la creiamo.
        answer = (
            Answer.objects.select_for_update()
            .filter(language=lang, question=q)
            .first()
        )
        if answer is None:
            answer = Answer(language=lang, question=q)

        del_ids = []
        for key, val in request.POST.items():
            if key.startswith("del_ex_") and (val or "").strip() == "1":
                try:
                    del_ids.append(int(key.split("_", 2)[2]))
                except ValueError:
                    pass

        updated_textareas = {}
        for key, val in request.POST.items():
            if not key.startswith("ex_"):
                continue
            try:
                _ex, ex_id, field = key.split("_", 2)
                ex_id = int(ex_id)
            except ValueError:
                continue
            if field == "textarea":
                updated_textareas[ex_id] = (val or "").strip()

        prefix = f"newex_{q.id}_"
        buckets = {}
        for key, val in request.POST.items():
            if not key.startswith(prefix):
                continue
            remainder = key[len(prefix):]
            try:
                uid, field = remainder.rsplit("_", 1)
            except ValueError:
                continue
            if field not in FIELDS:
                continue
            buckets.setdefault(uid, {})[field] = (val or "").strip()



        if response_text == "yes":
            nonempty_count = 0
            existing_qs = Example.objects.filter(answer__language=lang, answer__question=q)
            existing_map = {ex.id: ex for ex in existing_qs}

            # Conta gli esempi esistenti validi (non eliminati)
            for ex_id, ex in existing_map.items():
                if ex_id in del_ids:
                    continue
                tx = updated_textareas.get(ex_id, (ex.textarea or "").strip())
                if tx:
                    nonempty_count += 1

            # Conta i nuovi esempi validi creati
            if buckets:
                for _uid, data in buckets.items():
                    if (data.get("textarea") or "").strip():
                        nonempty_count += 1

            if nonempty_count < 2:
                messages.error(request,
                               _t("If you answer YES, you must provide at least two examples with a non-empty text."))
                return redirect(f"{reverse('language_data', kwargs={'lang_id': lang.id})}#p-{param.id}")




        answer.response_text = response_text
        answer.comments = comments
        answer.save()
        saved_count += 1

        try:
            target_ids = {int(x) for x in request.POST.getlist(f"mot_{q.id}")}
        except ValueError:
            target_ids = set()
        allowed_ids = set(
            QuestionAllowedMotivation.objects.filter(question=q).values_list("motivation_id", flat=True)
        )
        target_ids &= allowed_ids
        current_ids = set(AnswerMotivation.objects.filter(answer=answer).values_list("motivation_id", flat=True))
        to_add = target_ids - current_ids
        to_del = current_ids - target_ids
        if to_add:
            AnswerMotivation.objects.bulk_create(
                [AnswerMotivation(answer=answer, motivation_id=mid) for mid in to_add],
                ignore_conflicts=True,
            )
        if to_del:
            AnswerMotivation.objects.filter(answer=answer, motivation_id__in=to_del).delete()

        if del_ids:
            Example.objects.filter(answer=answer, id__in=del_ids).delete()

        for key, val in request.POST.items():
            if not key.startswith("ex_"):
                continue
            try:
                _ex, ex_id, field = key.split("_", 2)
                ex_id = int(ex_id)
            except ValueError:
                continue
            if field not in FIELDS:
                continue
            cleaned = (val or "").strip()
            Example.objects.filter(id=ex_id, answer=answer).update(**{field: cleaned})

        if buckets:
            to_create = []
            for idx, (uid, data) in enumerate(buckets.items()):
                has_payload = any([
                    data.get("textarea"),
                    data.get("transliteration"),
                    data.get("gloss"),
                    data.get("translation"),
                    data.get("reference"),
                ])
                if not has_payload:
                    continue
                to_create.append(Example(
                    answer=answer,
                    number=str(idx + 1),
                    textarea=data.get("textarea", ""),
                    transliteration=data.get("transliteration", ""),
                    gloss=data.get("gloss", ""),
                    translation=data.get("translation", ""),
                    reference=data.get("reference", ""),
                ))
            if to_create:
                Example.objects.bulk_create(to_create, ignore_conflicts=True)

    action = (request.POST.get("action") or "save").strip().lower() 
    
    # is_complete è True solo se il numero di risposte salvate ("yes"/"no") è uguale al numero totale di domande del parametro.
    is_complete = (len(questions) > 0 and saved_count == len(questions))

    # Best-effort: aggiornamento del flag isolato in un savepoint, così un suo
    # fallimento (es. deadlock o integrity error transitorio) non perde le
    # risposte appena salvate, ma viene comunque tracciato a log.
    try:
        with transaction.atomic():
            if action == "next" or not is_complete:
                ParameterReviewFlag.objects.update_or_create(
                    language=lang, parameter=param, user=request.user, defaults={"flag": True}
                )
            else:
                ParameterReviewFlag.objects.filter(language=lang, parameter=param).update(flag=False)
    except Exception:
        logger.exception(
            "Failed to update ParameterReviewFlag (lang=%s, param=%s, user=%s)",
            lang.id, param.id, request.user.id,
        )


    missing_count = len(questions) - saved_count
    
    if missing_count > 0:
        messages.warning(request, _t(f"Warning: {missing_count} missing answers for parameter {param.id}."))
    else:
        messages.success(request, _t(f"Saved {saved_count} answers for parameter {param.id}."))
    next_param = (
        ParameterDef.objects.filter(is_active=True, position__gt=param.position).order_by("position").first()
    )
    target_id = next_param.id if next_param else param.id
    return redirect(f"{reverse('language_data', kwargs={'lang_id': lang.id})}#p-{target_id}")




@login_required
@require_http_methods(["POST"])
@transaction.atomic
def answer_save(request: HttpRequest, lang_id: str, question_id: str) -> HttpResponse | None:
    """Save a single answer with motivations and examples.

    Args:
        request: Current authenticated HTTP request.
        lang_id: Primary key of the target language.
        question_id: Primary key of the question being answered.

    Returns:
        Redirect response to the language data page or language list depending
        on validation and permission checks.
    """

    lang = get_object_or_404(Language, pk=lang_id)
    if not _check_language_access(request.user, lang):
        messages.error(request, _t("You don't have access to this language."))
        return redirect("language_list")

    question = get_object_or_404(Question, pk=question_id)

    # 1) Input di base
    response_text = (request.POST.get("response_text") or "").strip().lower()
    if response_text not in ("yes", "no"):
        messages.error(request, _t("Invalid answer value."))
        return redirect("language_data", lang_id=lang.id)

    comments = (request.POST.get("comments") or "").strip()

    # 2) Lock/crea Answer
    answer = (
        Answer.objects.select_for_update()
        .filter(language=lang, question=question)
        .first()
    )
    if answer and not answer.modifiable and not _is_admin(request.user):
        messages.error(request, _t("This answer is locked (waiting/approved)."))
        return redirect("language_data", lang_id=lang.id)
    if answer is None:
        answer = Answer(language=lang, question=question)

    # 3) Salva header
    answer.response_text = response_text
    answer.comments = comments
    answer.save()

    # 4) Motivazioni
    try:
        motivation_ids = {int(x) for x in request.POST.getlist("motivation_ids")}
    except ValueError:
        motivation_ids = set()

    allowed_ids = set(
        QuestionAllowedMotivation.objects.filter(question=question).values_list("motivation_id", flat=True)
    )
    motivation_ids &= allowed_ids

    current_ids = set(AnswerMotivation.objects.filter(answer=answer).values_list("motivation_id", flat=True))
    to_add = motivation_ids - current_ids
    to_del = current_ids - motivation_ids
    if to_add:
        AnswerMotivation.objects.bulk_create(
            [AnswerMotivation(answer=answer, motivation_id=mid) for mid in to_add],
            ignore_conflicts=True,
        )
    if to_del:
        AnswerMotivation.objects.filter(answer=answer, motivation_id__in=to_del).delete()

    # 5) ESEMPI: raccolta mutazioni + VALIDAZIONE + applicazione (singola domanda)
    FIELDS = {"number", "textarea", "transliteration", "gloss", "translation", "reference"}

    # 5.1) Raccolta delete
    del_ids = []
    for key, val in request.POST.items():
        if not key.startswith("del_ex_"):
            continue
        if (val or "").strip() != "1":
            continue
        try:
            del_ids.append(int(key.split("_", 2)[2]))
        except ValueError:
            pass

    # 5.2) Raccolta update delle textarea esistenti
    updated_textareas = {}
    for key, val in request.POST.items():
        if not key.startswith("ex_"):
            continue
        try:
            _ex, ex_id, field = key.split("_", 2)
            ex_id = int(ex_id)
        except ValueError:
            continue
        if field == "textarea":
            updated_textareas[ex_id] = (val or "").strip()

    # 5.3) Raccolta nuovi esempi
    prefix = f"newex_{question.id}_"
    buckets = {}
    for key, val in request.POST.items():
        if not key.startswith(prefix):
            continue
        remainder = key[len(prefix):]
        try:
            uid, field = remainder.rsplit("_", 1)
        except ValueError:
            continue
        if field not in FIELDS:
            continue
        buckets.setdefault(uid, {})[field] = (val or "").strip()

    # 5.4) VALIDAZIONE per YES
    if response_text == "yes":

        if del_ids:
            Example.objects.filter(answer=answer, id__in=del_ids).delete()

        for key, val in request.POST.items():
            if not key.startswith("ex_"):
                continue
            try:
                _ex, ex_id, field = key.split("_", 2)
                ex_id = int(ex_id)
            except ValueError:
                continue
            if field not in FIELDS:
                continue
            cleaned = (val or "").strip()
            Example.objects.filter(id=ex_id, answer=answer).update(**{field: cleaned})

        if buckets:
            to_create = []
            for idx, (uid, data) in enumerate(buckets.items()):
                has_payload = any([
                    data.get("textarea"),
                    data.get("transliteration"),
                    data.get("gloss"),
                    data.get("translation"),
                    data.get("reference"),
                ])

                if not has_payload:
                    continue
                to_create.append(Example(
                    answer=answer,
                    num=str(idx + 1),
                    textarea=data.get("textarea", ""),
                    transliteration=data.get("transliteration", ""),
                    gloss=data.get("gloss", ""),
                    translation=data.get("translation", ""),
                    reference=data.get("reference", ""),
                ))
            if to_create:
                Example.objects.bulk_create(to_create, ignore_conflicts=True)

        messages.success(request, _t("Answer saved."))
        return redirect(f"{reverse('language_data', kwargs={'lang_id': lang.id})}#p-{question.parameter_id}")





@login_required
def language_save_instructions(request: HttpRequest, lang_id: str) -> HttpResponse:
    """Placeholder endpoint for saving language instructions.

    Args:
        request: Current authenticated HTTP request.
        lang_id: Primary key of the language.

    Returns:
        Redirect response with informative message.
    """
    if not _is_admin(request.user):
        messages.error(request, _t("You are not allowed to perform this action."))
        return redirect("language_list")
    messages.info(request, _t("Instructions are not supported yet (no model to store them)."))
    return redirect("language_data", lang_id=lang_id)


# -----------------------
# Pagina DEBUG con diagnostica e run DAG
# -----------------------
@login_required
def language_debug(request: HttpRequest, lang_id: str) -> HttpResponse:
    """Render the debug page with diagnostics and parameter values.

    Args:
        request: Current authenticated HTTP request.
        lang_id: Primary key of the target language.

    Returns:
        Rendered debug page for authorized admins.
    """

    user = request.user
    lang = get_object_or_404(Language, pk=lang_id)

    # Accesso
    if not _check_language_access(user, lang):
        get_object_or_404(Language, pk="__deny__")
    if not _is_admin(user):
        get_object_or_404(Language, pk="__deny__")

    # Parametri attivi + domande
    params = (
        ParameterDef.objects
        .filter(is_active=True)
        .order_by("position", "id")
        .prefetch_related(Prefetch("questions", queryset=Question.objects.order_by("id")))
    )

    # Risposte per lingua
    answers = (
        Answer.objects
        .filter(language=lang)
        .select_related("question")
        .order_by("question__parameter__position", "question_id")
    )
    answers_by_qid = {a.question_id: a for a in answers}

    # Valori iniziali/finali + warning
    lps = (
        LanguageParameter.objects
        .filter(language=lang)
        .select_related("parameter", "eval")
    )
    init_by_pid, warni_by_pid = {}, {}
    final_by_pid, warnf_by_pid = {}, {}
    for lp in lps:
        pid = lp.parameter_id
        init_by_pid[pid] = (lp.value_orig or "")
        warni_by_pid[pid] = bool(lp.warning_orig)
        if getattr(lp, "eval", None):
            final_by_pid[pid] = (lp.eval.value_eval or "")
            warnf_by_pid[pid] = bool(lp.eval.warning_eval)
        else:
            final_by_pid[pid] = ""
            warnf_by_pid[pid] = False

    diag_rows = diagnostics_for_language(lang)
    cond_map = {}
    for d in diag_rows:
        pid = getattr(d, "param_id", None) if not isinstance(d, dict) else d.get("param_id")
        cond_true = getattr(d, "cond_true", None) if not isinstance(d, dict) else d.get("cond_true")
        if pid:
            cond_map[pid] = cond_true  # True / False / None

    # Costruzione righe tabella principale (con cond_true)
    rows = []
    for p in params:
        q_ids, q_ans = [], []
        for q in p.questions.all():
            q_ids.append(q.id)
            a = answers_by_qid.get(q.id)
            q_ans.append(a.response_text.upper() if (a and a.response_text in ("yes", "no")) else "")
        rows.append({
            "position": p.position,
            "param_id": p.id,
            "name": p.name or "",
            "questions": q_ids,
            "answers": q_ans,
            "initial":  (init_by_pid.get(p.id, "") or ""),
            "final":    (final_by_pid.get(p.id, "") or ""),
            "warn_init": bool(warni_by_pid.get(p.id, False)),
            "warn_final": bool(warnf_by_pid.get(p.id, False)),
            "cond": (p.implicational_condition or ""),
            "cond_true": cond_map.get(p.id, None),  
        })

    ctx = {
        "language": lang,
        "rows": rows,
        "is_admin": _is_admin(user),
    }
    return render(request, "languages/debug_parameters.html", ctx)


# -----------------------
# Azione: esecuzione DAG manuale (Forza Approvazione anche se parziale)
# -----------------------
@login_required
@require_POST
def language_run_dag(request: HttpRequest, lang_id: str) -> HttpResponse:
    """Force-approve answers for a language and execute DAG evaluation.

    Args:
        request: Current authenticated HTTP request.
        lang_id: Primary key of the target language.

    Returns:
        Redirect response to the debug page with success/error feedback.
    """
    if not _is_admin(request.user):
        messages.error(request, _t("You are not allowed to run the DAG."))
        return redirect("language_debug", lang_id=lang_id)

    lang = get_object_or_404(Language, pk=lang_id)

    try:
        with transaction.atomic():
            changed = Answer.objects.filter(
                language=lang
            ).exclude(
                status=AnswerStatus.APPROVED
            ).update(
                status=AnswerStatus.APPROVED, modifiable=False
            )

            if changed > 0:
                LanguageReview.objects.create(language=lang, decision="approve", created_by=request.user)

            report = run_dag_for_language(lang_id)
    except Exception:
        logger.exception("language_run_dag failed for lang_id=%s; rolling back approval", lang_id)
        messages.error(request, _t("DAG failed; the forced approval was rolled back. Please retry or contact an administrator."))
        return redirect("language_debug", lang_id=lang_id)

    msg = _t(
        "Forced approval on %(n)d answers. DAG completed: processed %(p)d, forced to zero %(fz)d, missing orig %(mo)d, warnings propagated %(wp)d."
    ) % {
        "n": changed,
        "p": len(report.processed or []),
        "fz": len(report.forced_zero or []),
        "mo": len(report.missing_orig or []),
        "wp": len(report.warnings_propagated or []),
    }
    if report.missing_orig:
        msg += " Missing: " + ", ".join(report.missing_orig[:8]) + ("…" if len(report.missing_orig) > 8 else "")
    if report.parse_errors:
        msg += " ParseErrors: " + ", ".join(f"{pid}" for (pid, _, _) in report.parse_errors[:6]) + ("…" if len(report.parse_errors) > 6 else "")
    messages.success(request, msg)

    return redirect("language_debug", lang_id=lang_id)


# -----------------------
# Flusso submit/approve/reject/reopen
# -----------------------
@login_required
@require_POST
def language_submit(request: HttpRequest, lang_id: str) -> HttpResponse:
    """Submit language answers for admin review.

    Args:
        request: Current authenticated HTTP request.
        lang_id: Primary key of the target language.

    Returns:
        Redirect response to the language data page.
    """

    lang = get_object_or_404(Language, pk=lang_id)
    if not _check_language_access(request.user, lang):
        messages.error(request, _t("You don't have access to this language."))
        return redirect("language_list")

    if not _all_questions_answered(lang):
        messages.warning(request, _t("You submitted with some unanswered questions. An admin will review before approval."))

    changed = Answer.objects.filter(language=lang).update(
        status=AnswerStatus.WAITING, modifiable=False
    )
    messages.success(request, _t(f"Submitted {changed} answers for approval."))
    return redirect("language_data", lang_id=lang.id)


@login_required
@require_POST
def language_approve(request: HttpRequest, lang_id: str) -> HttpResponse:
    """Approve submitted answers and run DAG when answers are complete.

    Args:
        request: Current authenticated HTTP request.
        lang_id: Primary key of the target language.

    Returns:
        Redirect response to language data or debug page depending on outcome.
    """
    if not _is_admin(request.user):
        messages.error(request, _t("You are not allowed to perform this action."))
        return redirect("language_list")

    lang = get_object_or_404(Language, pk=lang_id)

    # Regola: tutte risposte devono esserci
    if not _all_questions_answered(lang):
        messages.error(request, _t("Cannot approve: there are unanswered questions. Complete all answers before starting the DAG."))
        return redirect("language_data", lang_id=lang.id)

    try:
        with transaction.atomic():
            # Approva solo ciò che è in WAITING; resta modifiable=False
            changed = Answer.objects.filter(language=lang, status=AnswerStatus.WAITING).update(
                status=AnswerStatus.APPROVED, modifiable=False
            )

            LanguageReview.objects.create(language=lang, decision="approve", created_by=request.user)

            # Avvia il DAG: se solleva, l'intera approvazione viene annullata
            report = run_dag_for_language(lang_id)
    except Exception:
        logger.exception("language_approve failed for lang_id=%s; rolling back approval", lang_id)
        messages.error(request, _t("Approval failed and was rolled back. Please retry or contact an administrator."))
        return redirect("language_data", lang_id=lang.id)

    msg = _t(
        "Approved %(n)d answers. DAG: processed %(p)d, forced_to_zero %(fz)d, missing_orig %(mo)d, warnings %(wp)d."
    ) % {
        "n": changed,
        "p": len(report.processed or []),
        "fz": len(report.forced_zero or []),
        "mo": len(report.missing_orig or []),
        "wp": len(report.warnings_propagated or []),
    }
    if report.parse_errors:
        msg += " ParseErrors: " + ", ".join(f"{pid}" for (pid, _, _) in report.parse_errors[:6])
    messages.success(request, msg)

    return redirect("language_debug", lang_id=lang.id)


@login_required
@require_POST
def language_reject(request: HttpRequest, lang_id: str) -> HttpResponse:
    """Reject a language submission and record the admin review.

    Args:
        request: Current authenticated HTTP request.
        lang_id: Primary key of the target language.

    Returns:
        Redirect response to the language data page.
    """

    if not _is_admin(request.user):
        messages.error(request, _t("You are not allowed to perform this action."))
        return redirect("language_list")

    lang = get_object_or_404(Language, pk=lang_id)
    message = (request.POST.get("message") or "").strip()

    with transaction.atomic():
        # Porta tutte le answers allo stato REJECTED e riapri editing
        changed = Answer.objects.filter(language=lang).update(
            status=AnswerStatus.REJECTED, modifiable=False
        )
        LanguageReview.objects.create(language=lang, decision="reject", message=message, created_by=request.user)

    if changed == 0:
        messages.info(request, _t("Nothing to reject."))
    else:
        messages.success(request, _t(f"Rejected submission. {changed} answers are editable again."))
    return redirect("language_data", lang_id=lang.id)


@login_required
@require_POST
def language_reopen(request: HttpRequest, lang_id: str) -> HttpResponse:
    """Reopen rejected answers by moving them back to pending.

    Args:
        request: Current authenticated HTTP request.
        lang_id: Primary key of the target language.

    Returns:
        Redirect response to the language data page.
    """

    lang = get_object_or_404(Language, pk=lang_id)
    if not _check_language_access(request.user, lang):
        messages.error(request, _t("You don't have access to this language."))
        return redirect("language_list")

    changed = Answer.objects.filter(language=lang, status=AnswerStatus.REJECTED).update(
        status=AnswerStatus.PENDING, modifiable=True
    )
    if changed == 0:
        messages.info(request, _t("Nothing to reopen."))
    else:
        messages.success(request, _t(f"Reopened: {changed} answers set to pending."))
    return redirect("language_data", lang_id=lang.id)




# === API per review flag (riuso di ParameterReviewFlag) =================

@login_required
@require_http_methods(["GET"])
def review_flags_list(request: HttpRequest, lang_id: str) -> JsonResponse:
    """Return the current user's active review flags for a language.

    Args:
        request: Current authenticated HTTP request.
        lang_id: Primary key of the target language.

    Returns:
        JSON payload with the list of flagged parameter IDs.
    """
    lang = get_object_or_404(Language, pk=lang_id)
    if not _check_language_access(request.user, lang):
        return JsonResponse({"flags": []})
    flags = list(
        ParameterReviewFlag.objects.filter(language=lang, flag=True)
        .values_list("parameter_id", flat=True)
    )
    return JsonResponse({"flags": flags})


@login_required
@require_http_methods(["POST"])
def toggle_review_flag(request: HttpRequest, lang_id: str, param_id: str) -> JsonResponse:
    """Set or clear a review flag for the current user and parameter.

    Args:
        request: Current authenticated HTTP request.
        lang_id: Primary key of the target language.
        param_id: Primary key of the target parameter.

    Returns:
        JSON payload with operation result and resulting flag state.
    """
    lang = get_object_or_404(Language, pk=lang_id)
    if not _check_language_access(request.user, lang):
        return JsonResponse({"ok": False, "error": "forbidden"}, status=403)
    param = get_object_or_404(ParameterDef, pk=param_id, is_active=True)
    flag = (request.POST.get("flag") or "").strip() == "1"
    ParameterReviewFlag.objects.update_or_create(
        language=lang, parameter=param, user=request.user, defaults={"flag": flag}
    )
    return JsonResponse({"ok": True, "flag": flag})


       
from datetime import datetime, time
from django.utils import timezone
from django.db.models import Max, Q



@login_required
@require_http_methods(["GET", "POST"])
def language_list_export_xlsx(request: HttpRequest) -> HttpResponse:
    q = (request.GET.get("q") or "").strip()
    user = request.user
    is_admin = _is_admin(user)

    qs = (
        Language.objects
        .select_related("assigned_user")
        .prefetch_related("users")
        .annotate(last_change=Max("answers__updated_at"))
        .order_by("position")
    )

    # 1. Gestione Selezione (POST) - Ha la priorità assoluta
    if request.method == "POST":
        lang_ids_str = request.POST.get("lang_ids", "")
        if lang_ids_str:
            selected_ids = [x.strip() for x in lang_ids_str.split(",") if x.strip()]
            if selected_ids:
                qs = qs.filter(id__in=selected_ids)
    
    # 2. Gestione Ricerca (GET) - Solo se non stiamo esportando selezionati
    else:
        if q:
            qs = qs.filter(
                Q(id__icontains=q) |
                Q(name_full__icontains=q) |
                Q(top_level_family__icontains=q) |
                Q(family__icontains=q) |
                Q(grp__icontains=q) |
                Q(isocode__icontains=q) |
                Q(glottocode__icontains=q) |
                Q(source__icontains=q) |
                Q(location__icontains=q)
            )

    if not is_admin:
        qs = qs.filter(Q(assigned_user=user) | Q(users=user)).distinct()

    def _xlsx_sanitize(v):
        if v is None:
            return ""
        if isinstance(v, bool):
            return "Yes" if v else "No"
        if isinstance(v, datetime):
            if timezone.is_aware(v):
                v = timezone.localtime(v)
            return v.replace(tzinfo=None)
        if isinstance(v, time):
            return v.replace(tzinfo=None) if v.tzinfo else v
        return v

    def _get_assigned_names(lang):
        users = list(lang.users.all())
        if users:
            return ", ".join(
                f"{(u.name or '').strip()} {(u.surname or '').strip()}".strip() or (u.email or "") for u in users)
        if lang.assigned_user:
            u = lang.assigned_user
            return f"{(u.name or '').strip()} {(u.surname or '').strip()}".strip() or (u.email or "")
        return ""

    def _get_assigned_emails(lang):
        users = list(lang.users.all())
        if users:
            return ", ".join(u.email for u in users if u.email)
        if lang.assigned_user:
            return lang.assigned_user.email or ""
        return ""

    HEADERS = [
            "Name",
            "ID",
            "Top-level family",
            "Family",
            "Group",
            "ISO code",
            "Glottocode",
            "Location",
            "Latitude",
            "Longitude",
            "Supervisor",
            "Informant",
            "Historical",
            "Source",
            "Assigned user",
            "Email",
            "Date last change",
        ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Languages"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # rows
    for L in qs.iterator(chunk_size=1000):
        row = [
            _xlsx_sanitize(L.name_full),
            _xlsx_sanitize(L.id),
            _xlsx_sanitize(L.top_level_family),
            _xlsx_sanitize(L.family),
            _xlsx_sanitize(L.grp),
            _xlsx_sanitize(L.isocode),
            _xlsx_sanitize(L.glottocode),
            _xlsx_sanitize(L.location),
            _xlsx_sanitize(L.latitude),
            _xlsx_sanitize(L.longitude),
            _xlsx_sanitize(L.supervisor),
            _xlsx_sanitize(L.informant),
            _xlsx_sanitize(L.historical_language),
            _xlsx_sanitize(L.source),
            _xlsx_sanitize(_get_assigned_names(L)),
            _xlsx_sanitize(_get_assigned_emails(L)),
            _xlsx_sanitize(L.last_change),
        ]
        ws.append(row)

    from openpyxl.utils import get_column_letter
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 8
        for cell in ws[col_letter]:
            val = cell.value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    ts = timezone.localtime(timezone.now()).strftime("%Y%m%d")
    filename = f"PCM_languages_{ts}.xlsx"
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


logger = logging.getLogger(__name__)  
def _run_import_language_from_excel_bg(tmp_path: str, original_name: str, user_id: int | None) -> None:
    """Run language import command in background and clean up temp file.

    Args:
        tmp_path: Absolute path to the temporary Excel file.
        original_name: Original uploaded filename used for logging.
        user_id: Optional ID of the user who triggered the import.

    Returns:
        None.
    """
    try:
        logger.info(
            "Starting Excel import in background: file=%s, user_id=%s",
            original_name,
            user_id,
        )
        # equiv. a: python manage.py import_language_from_excel --file <tmp_path>
        call_command("import_language_from_excel", file=tmp_path)
        logger.info("Excel import COMPLETED: file=%s", original_name)
    except Exception:
        logger.exception("Excel import FAILED: file=%s", original_name)
    finally:
        try:
            os.remove(tmp_path)
            logger.info("Temporary file deleted: %s", tmp_path)
        except OSError:
            logger.warning("Could not delete temporary file: %s", tmp_path)


@login_required
@require_http_methods(["GET", "POST"])
def language_import_excel(request: HttpRequest) -> HttpResponse:
    """Import languages from an uploaded Excel file.

    Args:
        request: Current authenticated HTTP request.

    Returns:
        Import page on GET, or redirect response with status messages on POST.
    """
    if not _is_admin(request.user):
        messages.error(request, _t("You are not allowed to perform this action."))
        return redirect("language_list")

    if request.method == "POST":
        upload = request.FILES.get("file")
        if not upload:
            messages.error(request, _t("You must select an Excel file to import."))
            return redirect("language_import_excel")

        filename = (upload.name or "").lower()
        if not filename.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
            messages.error(request, _t("Unsupported file type. Please upload an .xlsx file."))
            return redirect("language_import_excel")

        # Limite dimensione: protegge da timeout/OOM su file molto grandi.
        # Override via env DJANGO_EXCEL_IMPORT_MAX_MB (default: 25 MB).
        max_mb = int(os.environ.get("DJANGO_EXCEL_IMPORT_MAX_MB", "25"))
        max_bytes = max_mb * 1024 * 1024
        if upload.size and upload.size > max_bytes:
            messages.error(
                request,
                _t("File too large (max %(mb)d MB).") % {"mb": max_mb},
            )
            return redirect("language_import_excel")

        # Allowlist MIME per coerenza con l'estensione .xlsx*
        allowed_ct = {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel.sheet.macroEnabled.12",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.template",
            "application/vnd.ms-excel.template.macroEnabled.12",
            "application/octet-stream",  # alcuni browser inviano questo
        }
        if upload.content_type and upload.content_type not in allowed_ct:
            messages.error(request, _t("Unsupported file type. Please upload an .xlsx file."))
            return redirect("language_import_excel")

        tmp_path = None
        try:
            # Salvataggio del file in una posizione temporanea sul server
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                for chunk in upload.chunks():
                    tmp.write(chunk)
                tmp_path = tmp.name

            import io
            out = io.StringIO()
            call_command("import_language_from_excel", file=tmp_path, stdout=out)

            # LOG SOLO SU CONSOLE E MESSAGGIO PULITO PER L'UTENTE ---
            result_output = out.getvalue()
            if result_output:
                print(result_output)

            # Invia all'utente online esclusivamente il messaggio di successo generico
            messages.success(request, _t("Import completed successfully."))

        except Exception as e:
            logger.exception("Error during synchronous Excel import")
            messages.error(request, _t("Import failed: %(err)s") % {"err": str(e)})

        finally:
            if tmp_path and os.path.exists(tmp_path):
                os.remove(tmp_path)
        return redirect("language_list")

    return render(request, "languages/import_excel.html", {})



def _build_language_workbook(lang: Language, user: Any) -> tuple[Workbook, str]:
    """Build the Excel workbook for a single language export.

    The workbook structure is reused by both single-language and full ZIP
    exports. Admin users receive additional sheets.

    Args:
        lang: Language object to export.
        user: User requesting the export.

    Returns:
        A tuple ``(workbook, suffix)`` where suffix is ``full`` for admins and
        ``examples`` for non-admin users.
    """
    is_admin = _is_admin(user)

    # Parametri attivi
    params = (
        ParameterDef.objects
        .filter(is_active=True)
        .order_by("position", "id")
    )

    # Domande per parametro
    qs_by_param: dict[str, list[Question]] = {}
    for q in (
        Question.objects
        .select_related("parameter")
        .order_by("parameter__position", "id")
    ):
        qs_by_param.setdefault(q.parameter_id, []).append(q)

    # Risposte per lingua (indicizzate per question_id)
    answers = (
        Answer.objects
        .select_related("question")
        .filter(language_id=lang.id)
    )
    ans_by_qid = {a.question_id: a for a in answers}

    # Motivazioni in mappa id->label (per foglio Answers)
    mot_map = {
        m.id: getattr(m, "label", getattr(m, "text", ""))
        for m in Motivation.objects.all()
    }

    # --- Esempi per lingua: indicizzati per question_id e ordinati ---
    ex_by_qid: dict[str, list[Example]] = {}
    examples = (
        Example.objects
        .select_related("answer")
        .filter(answer__language_id=lang.id)
    )
    for ex in examples:
        qid = ex.answer.question_id
        ex_by_qid.setdefault(qid, []).append(ex)

    # Ordinamento per "number" numerico quando possibile
    for arr in ex_by_qid.values():
        def _as_int(v):
            try:
                return int(v)
            except Exception:
                return 10**9
        arr.sort(key=lambda e: _as_int(getattr(e, "number", "")))

    # --- Valori parametro per lingua: orig + eval (se presente) ---
    lps_qs = LanguageParameter.objects.filter(language=lang).select_related("parameter")
    if HAS_EVAL:
        lps_qs = lps_qs.select_related("eval")

    value_orig_by_pid: dict[str, str] = {}
    value_eval_by_pid: dict[str, str] = {}
    for lp in lps_qs:
        pid = lp.parameter_id
        value_orig_by_pid[pid] = (lp.value_orig or "")
        if getattr(lp, "eval", None):
            value_eval_by_pid[pid] = (getattr(lp.eval, "value_eval", None) or "")
        else:
            value_eval_by_pid[pid] = ""

    # === Workbook ===
    wb = Workbook()

    # Header fogli esistenti
    ans_header = [
        "Language ID", "Parameter Label", "Question ID", "Question",
        "Question status", "Answer", "Parameter value", "Motivation", "Comments",
    ]
    ex_header = [
        "Language ID", "Question ID", "Example #",
        "Example text", "Transliteration", "Gloss", "English translation", "Reference",
    ]

    # Header per foglio compatibile con import_language_from_excel
    upload_header = [
        "Language",
        "Parameter_Label",
        "Question_ID",
        "Question",
        "Question_Examples_YES",
        "Question_Intructions_Comments",
        "Language_Answer",
        "Language_Comments",
        "Language_Examples",
        "Language_Example_Gloss",
        "Language_Example_Translation",
        "Language_References",
    ]

    bold_white = Font(bold=True, color="FFFFFF")

    def _style_table(ws, name: str):
        max_col, max_row = ws.max_column, ws.max_row
        if max_row < 2:
            return
        ref = f"A1:{get_column_letter(max_col)}{max_row}"
        tbl = Table(displayName=name, ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(tbl)
        ws.freeze_panes = "A2"
        widths = (
            [14, 18, 12, 36, 18, 10, 16, 28, 26]  # Answers
            if name == "Answers"
            else [14, 12, 12, 36, 20, 20, 26, 24]  # Examples / Upload
        )
        for idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = w

    # === Foglio Examples: sempre presente (admin e user) ===
    ws_examples = wb.active
    ws_examples.title = "Examples"
    ws_examples.append(ex_header)
    for i in range(1, len(ex_header) + 1):
        ws_examples.cell(row=1, column=i).font = bold_white

    for p in params:
        for q in qs_by_param.get(p.id, []):
            for ex in ex_by_qid.get(q.id, []):
                ws_examples.append([
                    lang.id,
                    q.id,
                    getattr(ex, "number", ""),
                    getattr(ex, "textarea", ""),
                    getattr(ex, "transliteration", ""),
                    getattr(ex, "gloss", ""),
                    getattr(ex, "translation", ""),
                    getattr(ex, "reference", ""),
                ])
    _style_table(ws_examples, "Examples")

    # === Fogli aggiuntivi solo per admin ===
    if is_admin:
        # ----------------------------
        # Foglio Database_model (compatibile con l'IMPORT)
        # ----------------------------
        ws_upload = wb.create_sheet("Database_model", 0)  # primo foglio
        ws_upload.append(upload_header)
        for col_idx in range(1, len(upload_header) + 1):
            ws_upload.cell(row=1, column=col_idx).font = bold_white

        for p in params:
            p_label = p.id
            for q in qs_by_param.get(p.id, []):
                a = ans_by_qid.get(q.id)

                # Language_Answer: YES / NO / vuoto
                if a and a.response_text == "yes":
                    lang_answer = "YES"
                elif a and a.response_text == "no":
                    lang_answer = "NO"
                else:
                    lang_answer = ""

                lang_comments = getattr(a, "comments", "") if a else ""

                # Aggregazione esempi in celle multilinea
                ex_list = ex_by_qid.get(q.id, [])
                examples_lines: list[str] = []
                gloss_lines: list[str] = []
                transl_lines: list[str] = []
                ref_lines: list[str] = []

                for idx_ex, ex in enumerate(ex_list):
                    # Recupera il testo dell'esempio così com'è (con la numerazione manuale)
                    text = getattr(ex, "textarea", "") or ""

                    # Aggiunge il testo direttamente alla lista senza prefissi extra
                    examples_lines.append(text)

                    # Mantiene l'allineamento con glosse, traduzioni e riferimenti
                    gloss_lines.append(getattr(ex, "gloss", "") or "")
                    transl_lines.append(getattr(ex, "translation", "") or "")
                    ref_lines.append(getattr(ex, "reference", "") or "")

                cell_examples = "\n".join(examples_lines) if examples_lines else ""
                cell_gloss = "\n".join(gloss_lines) if gloss_lines else ""
                cell_transl = "\n".join(transl_lines) if transl_lines else ""
                cell_refs = "\n".join(ref_lines) if ref_lines else ""

                ws_upload.append([
                    # Language: il command di import usa name_full in colonna "Language"
                    lang.name_full,
                    p_label,
                    q.id,
                    getattr(q, "text", "") or "",
                    getattr(q, "example_yes", "") or "",
                    getattr(q, "instruction", "") or "",
                    lang_answer,
                    lang_comments,
                    cell_examples,
                    cell_gloss,
                    cell_transl,
                    cell_refs,
                ])

        _style_table(ws_upload, "Upload")

        # ----------------------------
        # Foglio Answers
        # ----------------------------
        ws_answers = wb.create_sheet("Answers", 1)  
        ws_answers.append(ans_header)
        for i in range(1, len(ans_header) + 1):
            ws_answers.cell(row=1, column=i).font = bold_white

        def _pretty_qc_from_status(status: str | None) -> str:
            s = (status or "").lower()
            if s == "approved":
                return "Done"
            if s in {"waiting_for_approval", "waiting"}:
                return "Needs review"
            return "Not compiled"

        for p in params:
            p_label = p.id
            for q in qs_by_param.get(p.id, []):
                a = ans_by_qid.get(q.id)
                param_value = (
                    value_eval_by_pid.get(q.parameter_id)
                    or value_orig_by_pid.get(q.parameter_id)
                    or ""
                )

                if a:
                    ids = list(
                        AnswerMotivation.objects
                        .filter(answer=a)
                        .values_list("motivation_id", flat=True)
                    )
                    mot_text = "; ".join(mot_map.get(i, str(i)) for i in ids)

                    ws_answers.append([
                        lang.id,
                        p_label,
                        q.id,
                        getattr(q, "text", ""),
                        _pretty_qc_from_status(getattr(a, "status", None)),
                        getattr(a, "response_text", ""),
                        param_value,
                        mot_text,
                        getattr(a, "comments", ""),
                    ])
                else:
                    ws_answers.append([
                        lang.id,
                        p_label,
                        q.id,
                        getattr(q, "text", ""),
                        "Not compiled",
                        "",
                        param_value,
                        "",
                        "",
                    ])
        _style_table(ws_answers, "Answers")

    suffix = "full" if is_admin else "examples"
    return wb, suffix




@login_required
def language_export_xlsx(request: HttpRequest, lang_id: str) -> HttpResponse:
    """Export one language workbook as XLSX.

    Args:
        request: Current authenticated HTTP request.
        lang_id: Primary key of the target language.

    Returns:
        XLSX file response for the selected language.

    Raises:
        Http404: If the language is not accessible for the current user.
    """
    lang = get_object_or_404(Language, pk=lang_id)
    if not _check_language_access(request.user, lang):
        raise Http404("Language not found")

    wb, suffix = _build_language_workbook(lang, request.user)

    ts = now().strftime("%Y%m%d")
    filename = f"PCM_{lang.id}_{suffix}_{ts}.xlsx"
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp



@login_required
@require_http_methods(["GET", "POST"]) 
def language_export_all_zip(request: HttpRequest) -> HttpResponse:
    """Export languages as a ZIP containing one XLSX per language.
    If specific IDs are provided via POST, exports only those.
    """
    if not _is_admin(request.user):
        messages.error(request, _t("You are not allowed to perform this action."))
        return redirect("language_list")

    # Lingue di default (tutte)
    langs = Language.objects.all().order_by("position", "id")

    # Se arrivano ID specifici, filtriamo il queryset
    if request.method == "POST":
        lang_ids_str = request.POST.get("lang_ids", "")
        if lang_ids_str:
            selected_ids = [x.strip() for x in lang_ids_str.split(",") if x.strip()]
            if selected_ids:
                langs = langs.filter(id__in=selected_ids)

    zip_buffer = io.BytesIO()
    ts = now().strftime("%Y%m%d")

    with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for lang in langs:
            wb, suffix = _build_language_workbook(lang, request.user)

            # serializza il singolo workbook in memoria
            xlsx_io = io.BytesIO()
            wb.save(xlsx_io)
            xlsx_io.seek(0)

            inner_name = f"PCM_{lang.id}_{suffix}_{ts}.xlsx"
            zf.writestr(inner_name, xlsx_io.getvalue())

    zip_buffer.seek(0)
    if request.method == "POST" and request.POST.get("lang_ids"):
        zip_filename = f"PCM_languages_selected_{ts}.zip"
    else:
        zip_filename = f"PCM_languages_full_{ts}.zip"

    resp = HttpResponse(zip_buffer.getvalue(), content_type="application/zip")
    resp["Content-Disposition"] = f'attachment; filename="{zip_filename}"'
    return resp


# ============================================================================
# MIGRATION BUNDLE: export completo per popolare il sito nuovo (FastAPI)
# ============================================================================

def _build_database_model_workbook(lang: Language) -> Workbook:
    """Workbook con il SOLO foglio Database_model per una lingua.

    Stesso identico formato del foglio Database_model prodotto da
    _build_language_workbook(), ma senza gli altri fogli: il bundle è
    già autoconsistente e non vogliamo duplicare schema/answers.
    """
    upload_header = [
        "Language", "Parameter_Label", "Question_ID", "Question",
        "Question_Examples_YES", "Question_Intructions_Comments",
        "Language_Answer", "Language_Comments",
        "Language_Motivations",
        "Language_Examples", "Language_Example_Transliteration",
        "Language_Example_Gloss", "Language_Example_Translation",
        "Language_References",
    ]
    bold_white = Font(bold=True, color="FFFFFF")

    params = ParameterDef.objects.filter(is_active=True).order_by("position", "id")
    qs_by_param: dict[str, list[Question]] = {}
    for q in Question.objects.order_by("parameter__position", "id"):
        qs_by_param.setdefault(q.parameter_id, []).append(q)

    answers = Answer.objects.select_related("question").filter(language_id=lang.id)
    ans_by_qid = {a.question_id: a for a in answers}

    # Motivazioni effettivamente associate a ogni risposta della lingua.
    # Mappa answer_id -> lista di Motivation.code, ordinata per code.
    motivations_by_answer_id: dict[int, list[str]] = {}
    am_qs = (
        AnswerMotivation.objects
        .filter(answer__language_id=lang.id)
        .select_related("motivation")
        .order_by("motivation__code")
    )
    for am in am_qs:
        code = (am.motivation.code or "") if am.motivation else ""
        if not code:
            continue
        motivations_by_answer_id.setdefault(am.answer_id, []).append(code)

    ex_by_qid: dict[str, list[Example]] = {}
    for ex in Example.objects.select_related("answer").filter(answer__language_id=lang.id):
        ex_by_qid.setdefault(ex.answer.question_id, []).append(ex)
    for arr in ex_by_qid.values():
        def _as_int(v):
            try: return int(v)
            except Exception: return 10**9
        arr.sort(key=lambda e: _as_int(getattr(e, "number", "")))

    wb = Workbook()
    ws = wb.active
    ws.title = "Database_model"
    ws.append(upload_header)
    for i in range(1, len(upload_header) + 1):
        ws.cell(row=1, column=i).font = bold_white

    for p in params:
        for q in qs_by_param.get(p.id, []):
            a = ans_by_qid.get(q.id)
            if a and a.response_text == "yes":
                lang_answer = "YES"
            elif a and a.response_text == "no":
                lang_answer = "NO"
            else:
                lang_answer = ""
            lang_comments = (a.comments or "") if a else ""

            # Codici delle motivazioni selezionate per questa risposta,
            # comma-separated su una sola riga (es. "MOT01, MOT07").
            mot_codes = motivations_by_answer_id.get(a.id, []) if a else []
            cell_motivations = ", ".join(mot_codes)

            ex_list = ex_by_qid.get(q.id, [])
            cell_ex = "\n".join((ex.textarea or "") for ex in ex_list)
            cell_tl = "\n".join((ex.transliteration or "") for ex in ex_list)
            cell_gl = "\n".join((ex.gloss or "") for ex in ex_list)
            cell_tr = "\n".join((ex.translation or "") for ex in ex_list)
            cell_rf = "\n".join((ex.reference or "") for ex in ex_list)

            ws.append([
                lang.name_full, p.id, q.id, q.text or "",
                q.example_yes or "", q.instruction or "",
                lang_answer, lang_comments,
                cell_motivations,
                cell_ex, cell_tl, cell_gl, cell_tr, cell_rf,
            ])
    ws.freeze_panes = "A2"
    return wb


def _build_languages_metadata_workbook() -> Workbook:
    """00_languages.xlsx: anagrafica completa con taxonomy."""
    headers = [
        "ID", "Name", "Position", "Top-level family", "Family", "Group",
        "ISO code", "Glottocode", "Location", "Latitude", "Longitude",
        "Supervisor", "Informant", "Historical", "Source",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Languages"
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
    for L in Language.objects.all().order_by("position", "id"):
        ws.append([
            L.id, L.name_full, L.position,
            L.top_level_family or "", L.family or "", L.grp or "",
            L.isocode or "", L.glottocode or "", L.location or "",
            float(L.latitude) if L.latitude is not None else "",
            float(L.longitude) if L.longitude is not None else "",
            L.supervisor or "", L.informant or "",
            "Yes" if L.historical_language else "No",
            L.source or "",
        ])
    ws.freeze_panes = "A2"
    return wb


def _build_motivations_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Motivations"
    ws.append(["ID", "Code", "Label"])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
    for m in Motivation.objects.all().order_by("code"):
        ws.append([m.id, m.code or "", m.label or ""])
    ws.freeze_panes = "A2"
    return wb


def _build_parameters_workbook() -> Workbook:
    headers = [
        "ID", "Position", "Name", "Schema", "Type", "Level",
        "Short Description", "Long Description",
        "Implicational Condition", "Explanation of Implicational Condition",
        "Is Active",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Parameters"
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
    for p in ParameterDef.objects.all().order_by("position", "id"):
        ws.append([
            p.id, p.position, p.name or "",
            p.schema or "", p.param_type or "", p.level_of_comparison or "",
            p.short_description or "", p.long_description or "",
            p.implicational_condition or "",
            p.description_of_the_implicational_condition or "",
            "Yes" if p.is_active else "No",
        ])
    ws.freeze_panes = "A2"
    return wb


def _build_questions_workbook() -> Workbook:
    headers = [
        "ID", "Parameter ID", "Text", "Template Type",
        "Instruction", "Instruction YES", "Instruction NO",
        "Example YES", "Help Info",
        "Is Stop Question", "Is Active",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
    # Nota: nel vecchio modello Question non ha is_active. Lo forziamo a Yes.
    for q in Question.objects.all().order_by("parameter_id", "id"):
        ws.append([
            q.id, q.parameter_id, q.text or "",
            q.template_type or "",
            q.instruction or "", q.instruction_yes or "", q.instruction_no or "",
            q.example_yes or "", q.help_info or "",
            "Yes" if getattr(q, "is_stop_question", False) else "No",
            "Yes",
        ])
    ws.freeze_panes = "A2"
    return wb


def _build_qam_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "QuestionAllowedMotivations"
    ws.append(["Question ID", "Motivation Code"])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
    rows = (
        QuestionAllowedMotivation.objects
        .select_related("motivation")
        .order_by("question_id", "position", "id")
    )
    for qam in rows:
        ws.append([qam.question_id, qam.motivation.code if qam.motivation else ""])
    ws.freeze_panes = "A2"
    return wb


def _build_glossary_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Glossary"
    ws.append(["Word", "Description"])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
    for g in Glossary.objects.all().order_by("word"):
        ws.append([g.word, g.description or ""])
    ws.freeze_panes = "A2"
    return wb


def _build_unsure_flags_workbook() -> Workbook:
    """Aggrega ParameterReviewFlag per (lang, param): se almeno un utente
    ha flag=True, esporta una riga. Nel sito nuovo diventerà is_unsure=True."""
    wb = Workbook()
    ws = wb.active
    ws.title = "UnsureFlags"
    ws.append(["Language ID", "Parameter ID"])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
    pairs = (
        ParameterReviewFlag.objects
        .filter(flag=True)
        .values_list("language_id", "parameter_id")
        .distinct()
        .order_by("language_id", "parameter_id")
    )
    for lid, pid in pairs:
        ws.append([lid, pid])
    ws.freeze_panes = "A2"
    return wb


@login_required
@require_http_methods(["GET", "POST"])
def db_backup_upload(request: HttpRequest) -> HttpResponse:
    """Carica un file .dump dal PC dell'admin nel volume condiviso /app/_backups/.

    Il file viene salvato come `uploaded_<timestamp>.dump`. Da console del
    container `db` lo si ripristina con:
        pg_restore -U "$POSTGRES_USER" -d "$POSTGRES_DB" --clean --if-exists \\
                   /backups/uploaded_<timestamp>.dump
    """
    if not _is_admin(request.user):
        messages.error(request, _t("You are not allowed to perform this action."))
        return redirect("language_list")

    backup_dir = "/app/_backups"
    if not os.path.isdir(backup_dir):
        messages.error(request, "Cartella backup non disponibile (volume non montato).")
        return redirect("language_list")

    if request.method == "POST":
        upload = request.FILES.get("file")
        if not upload:
            messages.error(request, "Devi selezionare un file .dump.")
            return redirect("db_backup_upload")

        fname = (upload.name or "").lower()
        if not fname.endswith(".dump"):
            messages.error(request, "Estensione non valida. Atteso un file .dump prodotto da pg_dump.")
            return redirect("db_backup_upload")

        # Limite 200 MB (override via env var DJANGO_DUMP_UPLOAD_MAX_MB)
        max_mb = int(os.environ.get("DJANGO_DUMP_UPLOAD_MAX_MB", "200"))
        max_bytes = max_mb * 1024 * 1024
        if upload.size and upload.size > max_bytes:
            messages.error(request, f"File troppo grande (max {max_mb} MB).")
            return redirect("db_backup_upload")

        ts = now().strftime("%Y%m%d_%H%M%S")
        dest = os.path.join(backup_dir, f"uploaded_{ts}.dump")
        try:
            with open(dest, "wb") as f:
                for chunk in upload.chunks():
                    f.write(chunk)
        except Exception as e:
            messages.error(request, f"Errore durante l'upload: {e}")
            return redirect("db_backup_upload")

        size_mb = os.path.getsize(dest) / (1024 * 1024)
        messages.success(
            request,
            mark_safe(
                f"Upload completato. File salvato come <code>uploaded_{ts}.dump</code> "
                f"({size_mb:.2f} MB). Per ripristinarlo, dalla console del container "
                f"<code>db</code> esegui:<br>"
                f"<code>pg_restore -U \"$POSTGRES_USER\" -d \"$POSTGRES_DB\" "
                f"--clean --if-exists /backups/uploaded_{ts}.dump</code>"
            ),
        )
        return redirect("db_backup_upload")

    # GET: mostra il form e la lista dei dump già presenti
    existing = []
    try:
        for path in sorted(
            glob.glob(os.path.join(backup_dir, "*.dump")),
            key=os.path.getmtime,
            reverse=True,
        ):
            stat = os.stat(path)
            existing.append({
                "name": os.path.basename(path),
                "size_mb": stat.st_size / (1024 * 1024),
                "mtime": datetime.fromtimestamp(stat.st_mtime),
            })
    except Exception:
        pass

    return render(
        request,
        "languages/db_backup_upload.html",
        {"existing": existing},
    )


@login_required
@require_http_methods(["GET"])
def db_backup_download(request: HttpRequest) -> HttpResponse:
    """Scarica il file .dump più recente dal volume condiviso /app/_backups (read-only).

    Il file viene prodotto da `pg_dump` lanciato dentro il container `db`
    sul path `/backups/*.dump`. Solo admin.
    """
    if not _is_admin(request.user):
        messages.error(request, _t("You are not allowed to perform this action."))
        return redirect("language_list")

    backup_dir = "/app/_backups"
    if not os.path.isdir(backup_dir):
        messages.error(request, "Cartella backup non disponibile (volume non montato).")
        return redirect("language_list")

    files = sorted(
        glob.glob(os.path.join(backup_dir, "*.dump")),
        key=os.path.getmtime,
        reverse=True,
    )
    if not files:
        messages.error(
            request,
            "Nessun file .dump trovato in /app/_backups. "
            "Lancia prima `pg_dump` dal container db verso /backups/.",
        )
        return redirect("language_list")

    latest = files[0]
    fname = os.path.basename(latest)
    return FileResponse(
        open(latest, "rb"),
        as_attachment=True,
        filename=fname,
        content_type="application/octet-stream",
    )


@login_required
@require_http_methods(["GET"])
def parameter_change_log_export_xlsx(request: HttpRequest) -> HttpResponse:
    """Esporta tutti i ParameterChangeLog in un singolo .xlsx (solo admin)."""
    if not _is_admin(request.user):
        messages.error(request, _t("You are not allowed to perform this action."))
        return redirect("language_list")

    headers = [
        "ID", "Changed at", "Parameter ID", "Parameter Name",
        "Changed by (email)", "Changed by (name)",
        "Recap", "Diff (JSON)",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "ParameterChangeLog"
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")

    rows = (
        ParameterChangeLog.objects
        .select_related("parameter", "changed_by")
        .order_by("-changed_at", "-id")
    )
    for log in rows:
        diff_str = ""
        if log.diff:
            try:
                diff_str = json.dumps(log.diff, ensure_ascii=False, sort_keys=True)
            except Exception:
                diff_str = str(log.diff)
        param_id = log.parameter_id or ""
        param_name = (log.parameter.name if log.parameter else "") or ""
        user_email = (log.changed_by.email if log.changed_by else "") or ""
        user_full = ""
        if log.changed_by:
            user_full = f"{log.changed_by.name or ''} {log.changed_by.surname or ''}".strip()
        changed_at = log.changed_at.replace(tzinfo=None) if log.changed_at else ""
        ws.append([
            log.id, changed_at, param_id, param_name,
            user_email, user_full, log.recap or "", diff_str,
        ])
    ws.freeze_panes = "A2"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    ts = now().strftime("%Y%m%d_%H%M%S")
    filename = f"parameter_change_log_{ts}.xlsx"
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@login_required
@require_http_methods(["GET"])
def language_export_migration_bundle(request: HttpRequest) -> HttpResponse:
    """Genera il Migration Bundle ZIP per popolare il sito nuovo (FastAPI).

    Contenuto:
        00_languages.xlsx
        01_motivations.xlsx
        02_parameters.xlsx
        03_questions.xlsx
        04_question_allowed_motivations.xlsx
        06_glossary.xlsx
        08_unsure_flags.xlsx
        data/<NomeLingua>.xlsx  (foglio Database_model per ogni lingua)

    Solo admin. Il file prodotto va caricato su:
        POST /api/admin/migration/import-bundle  (sito nuovo)
    """
    if not _is_admin(request.user):
        messages.error(request, _t("You are not allowed to perform this action."))
        return redirect("language_list")

    ts = now().strftime("%Y%m%d_%H%M%S")
    zip_buffer = io.BytesIO()

    def _save(wb):
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio.getvalue()

    with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("00_languages.xlsx", _save(_build_languages_metadata_workbook()))
        zf.writestr("01_motivations.xlsx", _save(_build_motivations_workbook()))
        zf.writestr("02_parameters.xlsx", _save(_build_parameters_workbook()))
        zf.writestr("03_questions.xlsx", _save(_build_questions_workbook()))
        zf.writestr("04_question_allowed_motivations.xlsx", _save(_build_qam_workbook()))
        zf.writestr("06_glossary.xlsx", _save(_build_glossary_workbook()))
        zf.writestr("08_unsure_flags.xlsx", _save(_build_unsure_flags_workbook()))

        # Una xlsx Database_model per lingua. Filename basato su name_full
        # sanificato per essere safe nel ZIP.
        import re
        used_names = set()
        for lang in Language.objects.all().order_by("position", "id"):
            safe = re.sub(r"[^A-Za-z0-9._-]+", "_", lang.name_full or lang.id)
            base = safe or lang.id
            name = f"{base}.xlsx"
            i = 2
            while name in used_names:
                name = f"{base}_{i}.xlsx"
                i += 1
            used_names.add(name)
            zf.writestr(f"data/{name}", _save(_build_database_model_workbook(lang)))

    zip_buffer.seek(0)
    filename = f"PCM_migration_{ts}.zip"
    resp = HttpResponse(zip_buffer.getvalue(), content_type="application/zip")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp