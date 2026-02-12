
import os

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models.signals import post_save, post_delete

try:
    from openpyxl import load_workbook
    HAS_XLSX = True
except Exception:
    HAS_XLSX = False

# import dei modelli necessari
from core.models import (
    Language,
    ParameterDef,
    Question,
    Answer,
    AnswerStatus,
    Example,
)
from core.signals import answer_saved_recompute, answer_deleted_recompute
from core.services.param_consolidate import recompute_and_persist_language_parameter


def _coerce_str(x):
    if x is None:
        return None
    return str(x)


def parse_null(v):
    return None if v is None or str(v).strip() == "" else v


def _split_lines(cell_value):
    if cell_value is None:
        return []
    s = str(cell_value).replace("\r\n", "\n").replace("\r", "\n").strip()
    if not s:
        return []
    return [line.strip() for line in s.split("\n") if line.strip()]


# Il testo rimarrà integrale (es: "1. Lorem ipsum")
def _split_examples(cell_value):
    lines = _split_lines(cell_value)
    out = []
    for idx, line in enumerate(lines):
        num = str(idx + 1)
        text = line.strip()
        out.append((num, text))
    return out


class Command(BaseCommand):
    help = (
        "Importa Answer ed Example da un file Excel denormalizzato "
        "(es. Database_Chioggia.xlsx), per una singola lingua."
    )

    def add_arguments(self, parser):
        # --file: percorso dell'Excel denormalizzato
        parser.add_argument(
            "--file",
            required=True,
            help="Percorso del file Excel denormalizzato (es. data/Database_Chioggia.xlsx)",
        )
        # --language-name: opzionale, full name della lingua (colonna 'Language')
        parser.add_argument(
            "--language-name",
            dest="language_name",
            help="Full name della lingua (colonna 'Language'); se omesso viene dedotto dal file.",
        )

    def handle(self, *args, **options):
        path = options["file"]
        language_name_opt = options.get("language_name")

        if not HAS_XLSX:
            raise CommandError("openpyxl non disponibile; installalo per usare questo comando.")

        if not os.path.exists(path):
            raise CommandError(f"File non trovato: {path}")

        # apertura Excel
        wb = load_workbook(path, data_only=True)
        if "Database_model" in wb.sheetnames:
            ws = wb["Database_model"]
        else:
            ws = wb.worksheets[0]

        rows_iter = ws.iter_rows(values_only=True)

        # Header
        try:
            headers = next(rows_iter)
        except StopIteration:
            raise CommandError("File Excel vuoto.")

        headers = [(_coerce_str(h) or "").strip() for h in headers]
        idx = {h: i for i, h in enumerate(headers) if h}

        # controlliamo che ci siano le colonne minime
        required_cols = ["Language", "Parameter_Label", "Question_ID", "Language_Answer"]
        missing = [c for c in required_cols if c not in idx]
        if missing:
            raise CommandError(
                "Colonne obbligatorie mancanti nel file: " + ", ".join(missing)
            )

        # carichiamo tutte le righe in memoria e raccogliamo i valori di Language
        all_rows = []
        lang_values = set()

        for raw in rows_iter:
            row_dict = {}
            empty = True
            for h, i in idx.items():
                val = raw[i] if i < len(raw) else None
                if val not in (None, ""):
                    empty = False
                row_dict[h] = val
            if empty:
                continue
            lang_val = row_dict.get("Language")
            if lang_val:
                lang_values.add(str(lang_val).strip())
            all_rows.append(row_dict)

        if not all_rows:
            self.stdout.write(self.style.WARNING("Nessuna riga dati trovata nel file."))
            return

        # determinazione della lingua
        if language_name_opt:
            language_name = language_name_opt.strip()
            if lang_values and language_name not in lang_values:
                # se non combacia, segnalo subito
                self.stdout.write(
                    self.style.WARNING(
                        f"Valori trovati in colonna 'Language': {sorted(lang_values)}"
                    )
                )
                raise CommandError(
                    f"language_name={language_name!r} non coerente con i dati del file."
                )
        else:
            if len(lang_values) != 1:
                raise CommandError(
                    "Impossibile dedurre in modo univoco la lingua dal file. "
                    f"Valori trovati in 'Language': {sorted(lang_values)}"
                )
            language_name = next(iter(lang_values))

        # recupero della Language dal DB (per name_full)
        try:
            language = Language.objects.get(name_full__iexact=language_name)
        except Language.DoesNotExist:
            raise CommandError(
                f"Lingua con name_full={language_name!r} non trovata nel DB."
            )

        self.stdout.write(
            self.style.SUCCESS(
                f"Import per lingua: {language_name} (id={language.id}) dal file {path}"
            )
        )

        imported_answers = 0
        imported_examples = 0
        skipped_rows = 0

        # NEW: disabilitiamo i signal per evitare ricalcoli ridondanti durante l'import
        post_save.disconnect(answer_saved_recompute, sender=Answer)
        post_delete.disconnect(answer_deleted_recompute, sender=Answer)

        try:
            # import atomico per sicurezza
            with transaction.atomic():
                # STEP 1: Rimuoviamo tutte le Answer esistenti per questa lingua
                # (approccio "replace all" per garantire consistenza)
                old_answers = Answer.objects.filter(language=language)
                old_count = old_answers.count()
                if old_count > 0:
                    self.stdout.write(
                        self.style.WARNING(
                            f"Rimozione di {old_count} Answer esistenti per {language_name}..."
                        )
                    )
                    # La delete in cascata rimuoverà anche Example e AnswerMotivation
                    old_answers.delete()

                # STEP 2: Importiamo le nuove Answer dal file
                for row in all_rows:
                    # Filtra per lingua (in caso il file contenga più lingue)
                    lang_val = (row.get("Language") or "").strip()
                    if lang_val and lang_val != language_name:
                        continue

                    param_label = (_coerce_str(row.get("Parameter_Label")) or "").strip()
                    if not param_label:
                        self.stdout.write(self.style.WARNING(f"Riga saltata per Parameter_Label mancante: {row}"))
                        skipped_rows += 1
                        continue

                    # Trova il parametro esistente
                    try:
                        param = ParameterDef.objects.get(id=param_label)
                    except ParameterDef.DoesNotExist:
                        self.stdout.write(
                            self.style.WARNING(
                                f"Parametro sconosciuto Parameter_Label={param_label!r}; riga saltata."
                            )
                        )
                        skipped_rows += 1
                        continue

                    qid = (_coerce_str(row.get("Question_ID")) or "").strip()
                    if not qid:
                        skipped_rows += 1
                        continue

                    q_text = parse_null(row.get("Question"))
                    q_ex_yes = parse_null(row.get("Question_Examples_YES"))
                    q_instr = parse_null(row.get("Question_Intructions_Comments"))

                    # Cerchiamo la domanda ignorando maiuscole/minuscole (es: PCA_Qsa -> PCA_QSa)
                    question = Question.objects.filter(id__iexact=qid).first()

                    if not question:
                        # Se la domanda non esiste proprio nel DB, segnaliamo l'errore e saltiamo la riga
                        self.stdout.write(
                            self.style.ERROR(
                                f"ERRORE: La Question_ID '{qid}' non esiste nel database. "
                                f"Controlla il codice nell'Excel. Riga saltata."
                            )
                        )
                        skipped_rows += 1
                        continue

                    # Se la domanda esiste, verifichiamo che appartenga al parametro corretto (es: PCA)
                    if question.parameter_id != param.id:
                        self.stdout.write(
                            self.style.WARNING(
                                f"ATTENZIONE: La domanda '{qid}' nel DB è legata a {question.parameter_id}, "
                                f"mentre nell'Excel è sotto {param_label}. Riga saltata."
                            )
                        )
                        skipped_rows += 1
                        continue

                    # NEW: mappiamo Language_Answer -> "yes"/"no"
                    raw_ans = (_coerce_str(row.get("Language_Answer")) or "").strip().upper()
                    if raw_ans in ("YES", "Y"):
                        resp = "yes"
                    elif raw_ans in ("NO", "N"):
                        resp = "no"
                    else:
                        # --- INIZIO MODIFICA: Segnala risposta non valida ---
                        self.stdout.write(
                            self.style.WARNING(f"Riga saltata: risposta '{raw_ans}' non valida per la domanda {qid}"))
                        # --- FINE MODIFICA ---
                        skipped_rows += 1
                        continue

                    comments = parse_null(row.get("Language_Comments"))

                    # NEW: crea la nuova Answer (dato che abbiamo fatto delete all sopra)
                    answer = Answer.objects.create(
                        language=language,
                        question=question,
                        status=AnswerStatus.PENDING,
                        modifiable=True,
                        response_text=resp,
                        comments=comments,
                    )
                    imported_answers += 1

                    # NEW: costruiamo gli Example dalle colonne Language_Examples/Gloss/Translation/References
                    ex_main = _split_examples(row.get("Language_Examples"))
                    gloss_lines = _split_lines(row.get("Language_Example_Gloss"))
                    trans_lines = _split_lines(row.get("Language_Example_Translation"))
                    ref_lines = _split_lines(row.get("Language_References"))

                    for idx_ex, (num, text_ex) in enumerate(ex_main):
                        gloss = gloss_lines[idx_ex] if idx_ex < len(gloss_lines) else None
                        transl = (
                            trans_lines[idx_ex] if idx_ex < len(trans_lines) else None
                        )
                        ref = ref_lines[idx_ex] if idx_ex < len(ref_lines) else None

                        Example.objects.create(
                            answer=answer,
                            number=num,
                            textarea=text_ex,
                            gloss=parse_null(gloss),
                            translation=parse_null(transl),
                            transliteration=None,
                            reference=parse_null(ref),
                        )
                        imported_examples += 1

        finally:
            # NEW: riconnettiamo i signal
            post_save.connect(answer_saved_recompute, sender=Answer)
            post_delete.connect(answer_deleted_recompute, sender=Answer)

        # STEP 3: Ricalcoliamo tutti i LanguageParameter per questa lingua
        # (una sola volta alla fine, invece che ad ogni Answer)
        self.stdout.write(
            self.style.SUCCESS(
                f"Import completato. Answers creati: {imported_answers}, "
                f"Examples creati: {imported_examples}, righe saltate: {skipped_rows}."
            )
        )

        self.stdout.write("Ricalcolo dei parametri per la lingua...")
        params_updated = set()
        for answer in Answer.objects.filter(language=language).select_related('question'):
            param_id = answer.question.parameter_id
            if param_id not in params_updated:
                recompute_and_persist_language_parameter(language.id, param_id)
                params_updated.add(param_id)

        self.stdout.write(
            self.style.SUCCESS(
                f"Ricalcolati {len(params_updated)} parametri per {language_name}."
            )
        )

