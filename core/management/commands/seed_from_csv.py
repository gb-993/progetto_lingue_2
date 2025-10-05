# core/management/commands/seed_from_csv.py

from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils import timezone  # noqa: F401 (potrebbe servire in futuro)
import os, csv

# Dipendenza leggera per Excel
try:
    from openpyxl import load_workbook
    HAS_XLSX = True
except Exception:
    HAS_XLSX = False

from core.models import (
    User, Glossary, ParameterDef, Language, Question,
    LanguageParameter, Motivation, Answer, Example,
    AnswerMotivation, LanguageParameterEval, AnswerStatus
)

DATA_DIR = "data"  # directory con i file .xlsx e/o .csv


# ---------------------- Helpers parsing ----------------------

def parse_bool(v):
    if v is None:
        return None
    s = str(v).strip().lower()
    return s in ("1", "true", "t", "yes", "y", "si", "s")

def parse_null(v):
    return None if v is None or str(v).strip() == "" else v

def _coerce_str(x):
    # Normalizza valori Excel numerici/date in str quando necessario
    if x is None:
        return None
    return str(x)

# ---------------------- IO: Excel / CSV ----------------------

def read_xlsx_by_filename(base_name):
    """
    Legge data/<base>.xlsx se esiste e restituisce list[dict].
    Usa la PRIMA sheet del file; header = prima riga.
    """
    if not HAS_XLSX:
        return [], False

    # Consenti passaggi tipo "users.csv" -> cerca "users.xlsx"
    if base_name.lower().endswith(".csv"):
        base_name = base_name[:-4] + ".xlsx"
    elif not base_name.lower().endswith(".xlsx"):
        base_name = base_name + ".xlsx"

    path = os.path.join(DATA_DIR, base_name)
    if not os.path.exists(path):
        return [], False

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.worksheets[0]  # prima sheet
    rows_iter = ws.iter_rows(values_only=True)

    try:
        headers = next(rows_iter)
    except StopIteration:
        return [], True  # file vuoto ma trovato

    headers = [(_coerce_str(h) or "").strip() for h in headers]
    out = []
    for row in rows_iter:
        d = {}
        # Mappa per nome colonna; se righe più corte, riempi con None
        for i, h in enumerate(headers):
            if not h:
                continue
            val = row[i] if i < len(row) else None
            d[h] = val
        # ignora righe completamente vuote
        if any(v is not None and str(v).strip() != "" for v in d.values()):
            out.append(d)
    return out, True

def read_csv_by_filename(base_name):
    """
    Legge data/<base>.csv se esiste e restituisce list[dict].
    """
    if base_name.lower().endswith(".xlsx"):
        base_name = base_name[:-5] + ".csv"
    elif not base_name.lower().endswith(".csv"):
        base_name = base_name + ".csv"

    path = os.path.join(DATA_DIR, base_name)
    if not os.path.exists(path):
        return [], False

    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f)), True

def read_table(name):
    """
    Fallback: prova prima XLSX, poi CSV. Ritorna (rows, source).
    source ∈ {"xlsx","csv","none"}
    """
    rows, found = read_xlsx_by_filename(name)
    if found:
        return rows, "xlsx"
    rows, found = read_csv_by_filename(name)
    if found:
        return rows, "csv"
    return [], "none"

def _status_line(label, source):
    if source == "xlsx":
        return f"{label} ok (xlsx)"
    if source == "csv":
        return f"{label} ok (csv)"
    return f"{label} (nessun file trovato: saltato)"

# ---------------------- Comando ----------------------

class Command(BaseCommand):
    help = "Importa dati iniziali da Excel (.xlsx) in data/; fallback ai CSV se l'Excel non esiste. Idempotente."

    @transaction.atomic
    def handle(self, *args, **opts):
        # 1) USERS
        users, src = read_table("users")
        for r in users:
            email = _coerce_str(r.get("email", "")).strip().lower()
            if not email:
                continue
            user, created = User.objects.get_or_create(
                email=email,
                defaults={
                    "name": _coerce_str(r.get("name", "")) or "",
                    "surname": _coerce_str(r.get("surname", "")) or "",
                    "role": _coerce_str(r.get("role", "user")) or "user",
                    "is_active": parse_bool(r.get("is_active")) if r.get("is_active") is not None else True,
                    "is_staff": bool(parse_bool(r.get("is_staff"))),
                    "is_superuser": bool(parse_bool(r.get("is_superuser"))),
                },
            )
            pwd = parse_null(r.get("password"))
            if pwd:
                user.set_password(_coerce_str(pwd))
                user.save(update_fields=["password"])
        self.stdout.write(self.style.SUCCESS(_status_line("Users", src)))

        # 2) GLOSSARY
        rows, src = read_table("glossary")
        for r in rows:
            word = _coerce_str(r.get("word", "")).strip()
            if not word:
                continue
            Glossary.objects.update_or_create(
                word=word,
                defaults={"description": _coerce_str(r.get("description", "")) or ""},
            )
        self.stdout.write(self.style.SUCCESS(_status_line("Glossary", src)))

        # 3) PARAMETERS
        rows, src = read_table("parameters")
        for r in rows:
            pid = _coerce_str(r.get("id", "")).strip()
            if not pid:
                continue
            ParameterDef.objects.update_or_create(
                id=pid,
                defaults={
                    "name": _coerce_str(r.get("name", "")) or "",
                    "short_description": parse_null(_coerce_str(r.get("short_description"))),
                    "position": int(r.get("position")) if r.get("position") not in (None, "") else 0,
                    "is_active": bool(parse_bool(r.get("is_active", True))),
                    "implicational_condition": parse_null(_coerce_str(r.get("implicational_condition"))),
                    "warning_default": bool(parse_bool(r.get("warning_default", False))),
                },
            )
        self.stdout.write(self.style.SUCCESS(_status_line("ParameterDef", src)))

        # 4) LANGUAGES
        users_by_email = {u.email: u.id for u in User.objects.all()}
        rows, src = read_table("languages")
        for r in rows:
            lid = _coerce_str(r.get("id", "")).strip()
            if not lid:
                continue
            assigned_email = _coerce_str(r.get("assigned_user_email") or "").strip().lower()
            assigned = users_by_email.get(assigned_email)
            Language.objects.update_or_create(
                id=lid,
                defaults={
                    "name_full": _coerce_str(r.get("name_full", "")) or "",
                    "position": int(r.get("position")) if r.get("position") not in (None, "") else 0,
                    "grp": parse_null(_coerce_str(r.get("grp"))),
                    "isocode": parse_null(_coerce_str(r.get("isocode"))),
                    "glottocode": parse_null(_coerce_str(r.get("glottocode"))),
                    "informant": parse_null(_coerce_str(r.get("informant"))),
                    "supervisor": parse_null(_coerce_str(r.get("supervisor"))),
                    "assigned_user_id": assigned,
                },
            )
        self.stdout.write(self.style.SUCCESS(_status_line("Languages", src)))

        # 5) QUESTIONS
        rows, src = read_table("questions")
        for r in rows:
            qid = _coerce_str(r.get("id", "")).strip()
            if not qid:
                continue
            Question.objects.update_or_create(
                id=qid,
                defaults={
                    "parameter_id": _coerce_str(r.get("parameter_id", "")).strip(),
                    "text": _coerce_str(r.get("text", "")) or "",
                    "example_yes": parse_null(_coerce_str(r.get("example_yes"))),
                    "instruction": parse_null(_coerce_str(r.get("instruction"))),
                    "template_type": parse_null(_coerce_str(r.get("template_type"))),
                    "is_stop_question": bool(parse_bool(r.get("is_stop_question", False))),
                },
            )
        self.stdout.write(self.style.SUCCESS(_status_line("Questions", src)))

        # 6) LANGUAGE_PARAMETERS
        rows, src = read_table("language_parameters")
        for r in rows:
            lang_id = _coerce_str(r.get("language_id", "")).strip()
            par_id  = _coerce_str(r.get("parameter_id", "")).strip()
            if not lang_id or not par_id:
                continue
            LanguageParameter.objects.update_or_create(
                language_id=lang_id,
                parameter_id=par_id,
                defaults={
                    "value_orig": _coerce_str(r.get("value_orig", "")).strip(),  # '+'|'-'
                    "warning_orig": bool(parse_bool(r.get("warning_orig", False))),
                },
            )
        self.stdout.write(self.style.SUCCESS(_status_line("LanguageParameter", src)))

        # 7) MOTIVATIONS
        rows, src = read_table("motivations")
        for r in rows:
            code = _coerce_str(r.get("code", "")).strip()
            if not code:
                continue
            Motivation.objects.update_or_create(
                code=code,
                defaults={"label": _coerce_str(r.get("label", "")) or ""},
            )
        self.stdout.write(self.style.SUCCESS(_status_line("Motivations", src)))

        # --- OPZIONALI ---

        # ANSWERS
        ans_map = {}
        rows, src_answers = read_table("answers")
        for r in rows:
            lang = _coerce_str(r.get("language_id", "")).strip()
            qid  = _coerce_str(r.get("question_id", "")).strip()
            if not lang or not qid:
                continue
            ans, _ = Answer.objects.update_or_create(
                language_id=lang, question_id=qid,
                defaults={
                    "status": _coerce_str(r.get("status", "pending")),
                    "modifiable": bool(parse_bool(r.get("modifiable", True))),
                    "response_text": _coerce_str(r.get("response_text", "yes")),
                    "comments": parse_null(_coerce_str(r.get("comments"))),
                },
            )
            ans_map[f"{lang}|{qid}"] = ans.id
        if rows:
            self.stdout.write(self.style.SUCCESS(_status_line("Answers", src_answers)))

        # EXAMPLES
        rows, src_ex = read_table("examples")
        for r in rows:
            key = _coerce_str(r.get("answer_lookup", "")).strip()  # es. "ita|FGMQ_a"
            if not key:
                continue
            ans_id = ans_map.get(key)
            if not ans_id:
                continue
            Example.objects.update_or_create(
                answer_id=ans_id,
                number=_coerce_str(r.get("number", "")),
                defaults={
                    "textarea": parse_null(_coerce_str(r.get("textarea"))),
                    "gloss": parse_null(_coerce_str(r.get("gloss"))),
                    "translation": parse_null(_coerce_str(r.get("translation"))),
                    "transliteration": parse_null(_coerce_str(r.get("transliteration"))),
                    "reference": parse_null(_coerce_str(r.get("reference"))),
                },
            )
        if rows:
            self.stdout.write(self.style.SUCCESS(_status_line("Examples", src_ex)))

        # ANSWER MOTIVATIONS
        rows, src_am = read_table("answer_motivations")
        if rows:
            # cache motivazioni per codice
            mot_by_code = {m.code: m.id for m in Motivation.objects.all()}
        for r in rows:
            key = _coerce_str(r.get("answer_lookup", "")).strip()
            mot_code = _coerce_str(r.get("motivation_code", "")).strip()
            if not key or not mot_code:
                continue
            ans_id = ans_map.get(key)
            mot_id = mot_by_code.get(mot_code)
            if not ans_id or not mot_id:
                continue
            AnswerMotivation.objects.get_or_create(
                answer_id=ans_id,
                motivation_id=mot_id,
            )
        if rows:
            self.stdout.write(self.style.SUCCESS(_status_line("AnswerMotivations", src_am)))

        # LANGUAGE PARAM EVAL
        rows, src_eval = read_table("language_parameter_eval")
        for r in rows:
            lang = _coerce_str(r.get("language_id", "")).strip()
            par  = _coerce_str(r.get("parameter_id", "")).strip()
            if not lang or not par:
                continue
            try:
                lp = LanguageParameter.objects.get(language_id=lang, parameter_id=par)
            except LanguageParameter.DoesNotExist:
                continue
            LanguageParameterEval.objects.update_or_create(
                language_parameter=lp,
                defaults={
                    "value_eval": _coerce_str(r.get("value_eval", "")).strip(),  # '+','-','0'
                    "warning_eval": bool(parse_bool(r.get("warning_eval", False))),
                },
            )
        if rows:
            self.stdout.write(self.style.SUCCESS(_status_line("LanguageParameterEval", src_eval)))

        self.stdout.write(self.style.SUCCESS("Seed completato (Excel/CSV)."))
