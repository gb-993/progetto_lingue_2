
import os
import re

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

try:
    from openpyxl import load_workbook
    HAS_XLSX = True
except Exception:
    HAS_XLSX = False

from core.models import Motivation, Question, QuestionAllowedMotivation

DATA_DIR = "data"


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def _split_mot_list(cell) -> list[str]:
    """
    Accetta: multilinea Excel, oppure stringa con separatori ; o ,
    """
    if cell is None:
        return []
    raw = str(cell).replace("\r\n", "\n").replace("\r", "\n")
    # prima split per newline
    parts = []
    for line in raw.split("\n"):
        line = line.strip()
        if not line:
            continue
        # poi split anche per ; o ,
        chunk = re.split(r"[;,]", line)
        for c in chunk:
            c = _norm(c)
            if c:
                parts.append(c)
    # de-dup preservando ordine
    seen = set()
    out = []
    for p in parts:
        if p in seen:
            continue
        seen.add(p)
        out.append(p)
    return out


def _iter_rows_2col(xlsx_path: str):
    """
    Legge la prima sheet di un xlsx e ritorna tuple(col1, col2) per ogni riga non vuota.
    Supporta anche file SENZA header: assume che tutte le righe siano dati.
    Se trova un header riconoscibile, lo salta.
    """
    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    first = True
    for row in ws.iter_rows(values_only=True):
        a = row[0] if len(row) > 0 else None
        b = row[1] if len(row) > 1 else None

        if a is None and b is None:
            continue

        # se prima riga sembra header, saltala
        if first:
            first = False
            ha = _norm(str(a)) if a is not None else ""
            hb = _norm(str(b)) if b is not None else ""
            headerish = {ha.lower(), hb.lower()}
            if (
                "question" in headerish
                or "question_id" in headerish
                or "domanda" in headerish
                or "motivations" in headerish
                or "motivation" in headerish
                or "motivazioni" in headerish
            ):
                continue

        yield a, b


class Command(BaseCommand):
    help = (
        "Importa motivations da data/motivations.xlsx (2 colonne: question_label | motivations) "
        "e popola QuestionAllowedMotivation. Crea automaticamente code MOT###."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default="motivations.xlsx",
            help="Nome file in data/ (default: motivations.xlsx)",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Non scrive sul DB, stampa solo le statistiche previste.",
        )

    @transaction.atomic
    def handle(self, *args, **opts):
        if not HAS_XLSX:
            raise CommandError("openpyxl non disponibile. Installalo per usare questo comando.")

        fname = opts["file"]
        dry = bool(opts["dry_run"])

        path = os.path.join(DATA_DIR, fname)
        if not os.path.exists(path):
            raise CommandError(f"File non trovato: {path}")

        # 1) Parse file: question_key -> [mot_label...]
        # a stessa question può apparire su più righe -> accumulo tutte le motivazioni
        q_to_mots: dict[str, list[str]] = {}
        all_mot_labels: list[str] = []

        for q_cell, mot_cell in _iter_rows_2col(path):
            q_key = _norm(str(q_cell)) if q_cell is not None else ""
            if not q_key:
                continue

            mot_list = _split_mot_list(mot_cell)

            # NEW: inizializza lista se non presente
            if q_key not in q_to_mots:
                q_to_mots[q_key] = []

            # NEW: se la riga non ha motivazioni, non deve "azzerare" quelle già viste
            if not mot_list:
                continue

            # NEW: merge senza duplicati, preservando ordine
            existing = q_to_mots[q_key]
            seen = set(existing)
            for m in mot_list:
                m = _norm(m)
                if not m or m in seen:
                    continue
                seen.add(m)
                existing.append(m)

            all_mot_labels.extend(mot_list)

        # 2) Costruisci set motivazioni uniche (ordine stabile)
        uniq_labels = []
        seen = set()
        for lab in all_mot_labels:
            if lab in seen:
                continue
            seen.add(lab)
            uniq_labels.append(lab)

        # 3) Crea/aggiorna Motivation con code MOT###
        #    NOTA: il code è solo un identificatore tecnico; label è quello che conta in UI.
        #    Per rendere l’operazione idempotente e stabile:
        #    - se esiste già una Motivation con lo stesso label, la riuso
        #    - altrimenti creo una nuova con prossimo code disponibile
        existing_by_label = { _norm(m.label): m for m in Motivation.objects.all() }
        existing_codes = set(Motivation.objects.values_list("code", flat=True))

        def next_code(i: int) -> str:
            return f"MOT{i:03d}"

        created_mot = 0
        reused_mot = 0

        # trova prossimo indice libero
        idx = 1
        while next_code(idx) in existing_codes:
            idx += 1

        label_to_mot_id: dict[str, int] = {}

        for lab in uniq_labels:
            nlab = _norm(lab)
            if nlab in existing_by_label:
                m = existing_by_label[nlab]
                label_to_mot_id[nlab] = m.id
                reused_mot += 1
                continue

            code = next_code(idx)
            idx += 1
            while code in existing_codes:
                code = next_code(idx)
                idx += 1

            if not dry:
                m = Motivation.objects.create(code=code, label=nlab)
                existing_codes.add(code)
                label_to_mot_id[nlab] = m.id
            created_mot += 1

        # 4) Risolvi le Question e aggiorna i link QuestionAllowedMotivation
        #    “tutte e sole”: per ogni domanda, cancello i link esistenti e ricreo quelli dal file.
        linked = 0
        missing_questions = 0

        for q_key, mot_list in q_to_mots.items():
            # prova: pk (Question.id), fallback: text case-insensitive
            q_obj = Question.objects.filter(pk=q_key).first()
            if q_obj is None:
                q_obj = Question.objects.filter(text__iexact=q_key).first()

            if q_obj is None:
                missing_questions += 1
                continue

            if not dry:
                QuestionAllowedMotivation.objects.filter(question=q_obj).delete()

                to_create = []
                pos = 1
                for lab in mot_list:
                    mid = label_to_mot_id.get(_norm(lab))
                    if not mid:
                        continue
                    to_create.append(
                        QuestionAllowedMotivation(
                            question=q_obj,
                            motivation_id=mid,
                            position=pos,
                        )
                    )
                    pos += 1

                if to_create:
                    QuestionAllowedMotivation.objects.bulk_create(to_create, ignore_conflicts=True)

            linked += 1

        self.stdout.write(self.style.SUCCESS(f"Letto file: {path}"))
        self.stdout.write(self.style.SUCCESS(f"Domande nel file: {len(q_to_mots)}"))
        self.stdout.write(self.style.SUCCESS(f"Motivazioni uniche nel file: {len(uniq_labels)}"))
        self.stdout.write(self.style.SUCCESS(f"Motivation riusate (label già presenti): {reused_mot}"))
        self.stdout.write(self.style.SUCCESS(f"Motivation nuove create: {created_mot}"))
        self.stdout.write(self.style.SUCCESS(f"Domande linkate aggiornate: {linked}"))
        if missing_questions:
            self.stdout.write(self.style.WARNING(f"Domande NON trovate nel DB: {missing_questions}"))
        if dry:
            self.stdout.write(self.style.WARNING("DRY-RUN: nessuna scrittura su DB eseguita."))
